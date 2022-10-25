#auto-py-to-exe
import minimalmodbus
import time
from tkinter import *
from tkinter.ttk import *
from tkinter import ttk
import matplotlib
import datetime
from tkinter import messagebox
import os
from PIL import Image, ImageTk
from numpy import var
import openpyxl
#import threading

def rtime(*args):#烘豆程式的計時器
    rt1 = args[0]   #傳入開機或入豆時間做為總烘豆時數的起點
    rt2=time.time() #程式循環一次後會呼叫此時間程式作為當下時間
    total_sec = rt2 - rt1   #計算總時數
    time_min = total_sec // 60
    time_sec = round((total_sec % 60),1)
    roast_time = str(int(time_min)) +':'+str(time_sec)
    RoT_2.config(text=roast_time)   #顯示總烘焙時間
    return(total_sec)

def roast_state(state_r): # 設定烘焙階段
    #使用全域變數,讓狀態碼可以傳入溫度偵測回圈內
    global state_a
    if state_r =='CHARGE':
        canvas.delete("all")
        draw_panal()
        canvas_ss.delete("all")
        draw_panal_ss()

        bt_charge.config(state=DISABLED)#入豆
    elif state_r =='TP':
        bt_tp.config(state=DISABLED)#回溫點
    elif state_r =='DRYe':
        bt_drye.config(state=DISABLED)#脫水結束
    elif state_r =='GDp':
        bt_gdp.config(state=DISABLED)#金黃點
    elif state_r =='FCs':
        bt_fcs.config(state=DISABLED)#一爆
    elif state_r =='FCe':
        bt_fce.config(state=DISABLED)#一爆結束
    elif state_r =='SCs':
        bt_scs.config(state=DISABLED)#二爆
    elif state_r =='SCe':
        bt_sce.config(state=DISABLED)#二爆結束

    state_a = state_r

    return state_a

def save_data(time_data,bt_temperature_data,ror_bt_data,et_temperature_data,ror_et_data,entry_temperature_data,event_data,step_data):#儲存資料
    filename = filename_E.get() #檔名
    #print(bt_temperature_data)
    i = 0
    with open( filename +'.rxt','w') as f:
        f.write(roastdate_E.get())  #烘焙日期
        f.write('\n')
        f.write(prodname_E.get())   #產品名稱
        f.write('\n')
        f.write(gb_method_E.get())   #生豆處理法
        f.write('\n')
        f.write(gb_name_E.get())   #生豆名稱
        f.write('\n')
        f.write(gb_moisture_E.get())   #水分含量
        f.write('\n')
        f.write(gb_density_E.get())   #生豆密度
        f.write('\n')
        f.write(gb_weight_E.get())   #批次重量
        f.write('\n')
        f.write(machine_name_E.get())   #設備名稱
        f.write('\n')
        f.write(content_E.get())   #設備容量
        f.write('\n')
        f.write(energy_E.get())   #使用能源
        f.write('\n')
        f.write(operator_E.get())   #操作人員
        f.write('\n')
        f.write(weather_E.get())   #天氣溫度
        f.write('\n')
        f.write(cb_weight_E.get())   #熟豆重量
        f.write('\n')
        f.write(loss_weight_E.get())   #失重率
        f.write('\n')
        f.write(cb_barg_E.get())   #豆表色度
        f.write('\n')
        f.write(cb_parg_E.get())   #豆粉色度
        f.write('\n')
        f.write(cb_cupping_E.get())   #杯測風味
        f.write('\n')

        for i in range(len(time_data)-1):   #存檔順序 序號 :豆溫:豆溫ror:環境溫:環境溫ror:入風溫:時間:事件
            f.write(str(i+1)+':')
            f.write(str(bt_temperature_data[i]))
            f.write(':')
            f.write(str(ror_bt_data[i]))
            f.write(':')
            f.write(str(et_temperature_data[i]))
            f.write(':')
            f.write(str(ror_et_data[i]))
            f.write(':')
            f.write(str(entry_temperature_data[i]))
            f.write(':')
            f.write(str(round((time_data[i] / 0.5),1)))#原本資料是座標軸單位,所以需轉回時間(秒)
            f.write(':')
            f.write(str(event_data[i]))
            f.write('\n')
        try:
            f.write(str(step_data[0]))
            f.write('\n')
            f.write(str(step_data[1]))
            f.write('\n')
            f.write(str(step_data[2]))
            f.write('\n')
        except:
            messagebox.showinfo('information', '資料不齊全，無法計算階段百分比')
        #else:
        messagebox.showinfo('information', '存檔完成')
    return            

def clean_tree():#清除Tree資料表、資料欄位及相關按鈕復歸
    #****** 清除Tree資料表 ******
    tree_item = tree.get_children()[0];iids = [] #'I001'
    #print(tree_item)
    while tree.exists(tree_item) == True :
        if tree_item != '':
            iids.append(tree_item)
        tree_item = tree.next(tree_item)
        if tree_item == '' :
            break
        continue
    for iid in iids:
        tree.delete(iid)
    RoT_2.config(text='')   #清除烘豆時間欄位
    even_dev_p_l.config(text='')  #清除發展時間欄位
    even_dev_l.config(text='')  #清除發展比例欄位
    even_predryend_l.config(text='')  #清除預估脫水結束時間欄位
    #et_temp_ror_l.config(text='')   #清除環境溫ROR欄位
    et_temp_l.config(text='')   #清除環境溫R欄位
    bt_temp_ror_l.config(text='')   #清除豆溫ROR欄位
    bt_temp_l.config(text='')   #清除豆溫欄位        
    f1_agtron_L_S.config(text='')   #清除艾格狀指數預估欄位

    draw_panal()    #重繪繪圖座標系統
    draw_panal_ss(typess_no)    #重繪繪圖座標系統

    #將事件按鈕復歸
    bt_openmc.config(state='')#開機
    bt_records.config(state=DISABLED)#開始紀錄
    bt_charge.config(state=DISABLED)#入豆
    bt_tp.config(state=DISABLED)#回溫點
    bt_drye.config(state=DISABLED)#脫水結束
    bt_gdp.config(state=DISABLED)#金黃點
    bt_fcs.config(state='')#一爆
    bt_fce.config(state='')#一爆結束
    bt_scs.config(state='')#二爆
    bt_sce.config(state='')#二爆結束
    bt_drop.config(state='')#下豆
    bt_cleantr.config(state=DISABLED)#清除表格
    #將事件狀態發生的資訊欄位資料清空
    charge_E.delete(0,END)
    rtp_E.delete(0,END)
    enddry_E.delete(0,END)
    gp_E.delete(0,END)
    fc_E.delete(0,END)
    fcend_E.delete(0,END)
    sc_E.delete(0,END)
    scend_E.delete(0,END)
    drop_E.delete(0,END)
    #********************************************************
    var_rostype = IntVar()
    var_rostype.set(2)
    rostype_nd_1 = Radiobutton(frame4,text="北歐烘焙法",variable=var_rostype,value = 1,command=lambda:draw_panal_ss(1))
    rostype_nd_1.grid(row=0,column=0,columnspan=14,padx=5,pady=5)
    rostype_sr_2 = Radiobutton(frame4,text="Scott Rao_漸降式烘焙法",variable=var_rostype,value = 2,command=lambda:draw_panal_ss(2))
    rostype_sr_2.grid(row=0,column=15,columnspan=14,padx=5,pady=5)
    rostype_sn_3 = Radiobutton(frame4,text="小野善造_完全烘焙法",variable=var_rostype,value = 3,command=lambda:draw_panal_ss(3))
    rostype_sn_3.grid(row=15,column=0,columnspan=14,padx=5,pady=5)
    rostype_ss_4 = Radiobutton(frame4,text="全息烘焙法",variable=var_rostype,value = 4,command=lambda:draw_panal_ss(4))
    rostype_ss_4.grid(row=15,column=15,columnspan=14,padx=5,pady=5)

    messagebox.showinfo('information', '清除完成')
    return

def draw_panal():#重新產生畫布座標系統
    canvas.delete("all")    #清除繪圖畫面
    #第一座標Y軸
    canvas.create_line(36,0,36,440,width=2,fill='blue')#Y 第一座標軸    
    for i in range(20,420,80):
        canvas.create_line(36,i,46,i,width=2,fill='red')
        canvas.create_text(20,i,text=str((5-(i // 80))*50),fill='blue')
        if i > 20 :
            canvas.create_line(47,i,900,i,width=1,fill='#0f1', dash=(10,2))

    #第二座標Y軸
    canvas.create_line(900,0,900,440,width=2,fill='red')#Y 第二座標軸    
    for j in range(20,420,80):
        canvas.create_line(900,j,910,j,width=2,fill='red')
        canvas.create_text(920,j,text=str((5-(j // 80))*10),fill='red')
 
    #X座標軸
    canvas.create_line(36,420,900,420,width=2,fill='green')#X 時間軸    
    jup = 0    
    for k in range(36,876,30):
        jup += 1
        if jup == 2 or jup == 7 or jup == 12 or jup == 17 or jup == 22 or jup == 27:
            canvas.create_line(k,420,k,410,width=2,fill='red')
            canvas.create_text(k,430,text=str(jup-2),fill='blue')
        else:
            canvas.create_line(k,420,k,410,width=2,fill='green')

        if k > 36 :
            canvas.create_line(k,50,k,410,width=1,fill='#0f1', dash=(10,2)) 

    #********** 事件座標軸 **********
    canvas_event.delete("all")    #清除事件繪圖畫面
    #事件座標Y1軸
    canvas_event.create_line(36,10,36,120,width=2,fill='blue')#Y 第一座標軸    
    for i in range(10,110,20):
        canvas_event.create_line(36,i,46,i,width=2,fill='red')
        if i % 20 == 10:
            canvas_event.create_text(20,i,text=str((11-(i // 10))*5),fill='blue')
        if i > 10 :
            canvas_event.create_line(47,i,900,i,width=1,fill='#fac', dash=(10,2))
    canvas_event.create_text(20,110,text='0',fill='blue')
    canvas_event.create_text(36,122,text='壓差',fill='blue')

    #事件座標Y2軸
    canvas_event.create_line(900,10,900,120,width=2,fill='blue')#Y 第一座標軸    
    for i in range(10,110,20):
        canvas_event.create_line(900,i,890,i,width=2,fill='red')
        if i % 40 == 10:
            canvas_event.create_text(920,i,text=str((11-(i // 10))*2.5),fill='red')
        if i > 10 :
            canvas_event.create_line(920,i,900,i,width=1,fill='#fac', dash=(10,2))
    canvas_event.create_text(920,110,text='0',fill='red')
    canvas_event.create_text(900,122,text='瓦斯',fill='red')

    #事件X座標軸
    canvas_event.create_line(36,110,900,110,width=2,fill='green')#X 時間軸    
    jup = 0    
    for k in range(36,876,30):
        jup += 1
        if jup == 2 or jup == 7 or jup == 12 or jup == 17 or jup == 22 or jup == 27:
            canvas_event.create_line(k,100,k,110,width=2,fill='red')
            canvas_event.create_text(k,118,text=str(jup-2),fill='blue')
        else:
            canvas_event.create_line(k,100,k,110,width=2,fill='green')

        if k > 36 :
            canvas_event.create_line(k,10,k,110,width=1,fill='#fac', dash=(10,2)) 
    #階段
    dry_start = float(rostep_dry_start_E.get())#內定脫水起始溫度 
    dry_end = float(rostep_dry_end_E.get())#內定脫水結束溫度
    maillard_start = float(rostep_maillard_start_E.get())#內定梅納階段起始溫度
    maillard_end = float(rostep_maillard_end_E.get())#內定梅納階段結束溫度
    development_start = float(rostep_development_start_E.get())#內定完成起始溫度
    development_end = float(rostep_development_end_E.get())#內定完成結束溫度

    canvas.create_rectangle(36, (420-dry_start*1.6), 900, (420-dry_end*1.6),fill='lightgreen', stipple="gray25" )#脫水期activefill='lightgreen',activestipple="gray50"
    canvas.create_rectangle(36, (420-maillard_start*1.6), 900, (420-maillard_end*1.6),fill='#ff8000', stipple="gray25" )#梅納期skyblue
    canvas.create_rectangle(36, (420-development_start*1.6), 900, (420-development_end*1.6),fill='gray', stipple="gray25" )#發展期skyblue
    return 

def draw_panal_ss(*sstype):#重新產生畫布座標系統__Frame6 全息烘焙
    global typess_no
    canvas_ss.delete("all")    #清除繪圖畫面
    #第一座標Y軸
    canvas_ss.create_line(36,0,36,440,width=2,fill='blue')#Y 第一座標軸    
    for i in range(20,420,80):
        canvas_ss.create_line(36,i,46,i,width=2,fill='red')
        canvas_ss.create_text(20,i,text=str((5-(i // 80))*50),fill='blue')
        if i > 20 :
            canvas_ss.create_line(47,i,900,i,width=2,fill='#0f1', dash=(10,2))

    #第二座標Y軸
    canvas_ss.create_line(900,0,900,440,width=2,fill='red')#Y 第二座標軸    
    for j in range(20,420,80):
        canvas_ss.create_line(900,j,910,j,width=2,fill='red')
        canvas_ss.create_text(920,j,text=str((5-(j // 80))*10),fill='red')
 
    #X座標軸
    canvas_ss.create_line(36,420,900,420,width=2,fill='green')#X 時間軸    
    jup = 0    
    for k in range(36,876,30):
        jup += 1
        if jup == 2 or jup == 7 or jup == 12 or jup == 17 or jup == 22 or jup == 27:
            canvas_ss.create_line(k,420,k,410,width=2,fill='red')
            canvas_ss.create_text(k,430,text=str(jup-2),fill='blue')
        else:
            canvas_ss.create_line(k,420,k,410,width=2,fill='green')

        if k > 36 :
            canvas_ss.create_line(k,50,k,410,width=2,fill='#0f1', dash=(10,2)) 
    
    #繪製各階段的底圖
    if sstype[0] == 1: #北歐
        typess_no = 1
        canvas_ss.create_rectangle(66+22.5, (420-75*1.6), 66+30, (420-120*1.6),fill='lightgreen', stipple="gray50" )#回溫點區域
        canvas_ss.create_rectangle(66+285, (420-180*1.6), 66+300, (420-230*1.6),fill='brown', stipple="gray50" )#回溫點區域

        #print(sstype)
    elif sstype[0] == 2: #Scott Rao
        typess_no = 2
        #print(sstype)
        dry_start = float(rostep_dry_start_E.get())#內定脫水起始溫度 
        dry_end = float(rostep_dry_end_E.get())#內定脫水結束溫度
        maillard_start = float(rostep_maillard_start_E.get())#內定梅納階段起始溫度
        maillard_end = float(rostep_maillard_end_E.get())#內定梅納階段結束溫度
        development_start = float(rostep_development_start_E.get())#內定完成起始溫度
        development_end = float(rostep_development_end_E.get())#內定完成結束溫度
        canvas_ss.create_rectangle(36, (420-dry_start*1.6), 900, (420-dry_end*1.6),fill='lightgreen', stipple="gray50" )#脫水期activefill='lightgreen',activestipple="gray50"
        canvas_ss.create_rectangle(36, (420-maillard_start*1.6), 900, (420-maillard_end*1.6),fill='#ff8000', stipple="gray50" )#梅納期skyblue
        canvas_ss.create_rectangle(36, (420-development_start*1.6), 900, (420-development_end*1.6),fill='gray', stipple="gray50" )#發展期skyblue
    elif sstype[0] == 3: #小野繕造
        typess_no = 3
        #print(sstype)

    elif sstype[0] == 4:#繪製全息烘焙的預估底圖
        typess_no = 4
        t1=5.0 #t1時間
        t2=9.5 #t2時間
        tf=11 #FCs時間
        td=15 #Drop時間
        #print(sstype)
        canvas_ss.create_rectangle(66, (420-0*1.6), (66+t1*30), (420-230*1.6),fill='lightgreen', stipple="gray12" )#回溫點範圍
        canvas_ss.create_rectangle((66+t1*30), (420-0*1.6), (66+t2*30), (420-230*1.6),fill='blue', stipple="gray12" )#T0點範圍
        canvas_ss.create_rectangle((66+t2*30), (420-0*1.6), (66+tf*30), (420-230*1.6),fill='#ff8000', stipple="gray12" )#T1點範圍
        canvas_ss.create_rectangle((66+tf*30), (420-0*1.6), (66+td*30), (420-230*1.6),fill='red', stipple="gray12" )#T2點範圍
        #T0、T1、T2、FCs起始線

    else:
        typess_no = 0
        #print("未選擇烘焙法")

    return  

def temp_ror(state_arg):#主程式_溫度擷取及烘豆階段紀錄
    
    global time_data    #時間資料陣列
    global bt_temperature_data  #豆溫資料陣列
    global et_temperature_data  #環境溫資料陣列
    global entry_temperature_data  #入風溫資料陣列
    global event_data   #事件資料陣列
    global counter_data #程式執行次數計數器
    global ror_bt_data  #豆溫ROR資料陣列
    global ror_et_data  #環境溫ROR資料陣列
    global ET   #偵測到的環境溫
    global BT   #偵測到的豆溫
    global entry   #偵測到的入風溫   
    global BT_ror   #豆溫ROR
    global ET_ror   #環境溫ROR
    global state_a  #事件狀態變數
    global counter_data_ax  #時間座標變數
    global counter_data_temp    #控制程式每秒只讀取一次溫度資料
    global rt1  #入豆時間點
    global roast_time   #總烘豆時間
    global record_start_flag #開始記錄旗標
    global charge_flag  #入豆旗標
    global rtp_flag #回溫點旗標
    global dry_end_flag #脫水結束旗標
    global gp_flag #金黃點旗標        
    global fc_flag  #1爆點旗標
    global fc_end_flag  #1爆結束旗標
    global secondc_flag #2爆點旗標
    global secondc_end_flag #2爆結束旗標.
    global drop_flag    #下豆點旗標
    global pre_base_time #估算脫水結束的基礎時間
    global fc_start_time #一爆點起始時間
    global step_data    #階段時間資料
    global ss_t0_flag   #全息烘焙 T0 點
    global ss_t1_flag   #全息烘焙 T1 點
    global ss_t2_flag   #全息烘焙 T2 點
    global agtron_num   #艾格狀數值
    global agtron_temp  #艾格狀數值基礎溫度
    global t0_temp  #T0點溫度
    global t1_temp  #T1點溫度
    global t2_temp  #T2點溫度
    global typess_no #烘焙方式選擇
    global ep_y_0
    global ep_y_1
    global ef_y_0
    global ef_y_1

    rt1=time.time()#rt1 = 0
    state_a = state_arg

    #將開始按鈕隱藏,防止誤按
    bt_openmc.config(state=DISABLED)#開機
    bt_recorde.config(state='')#結束記錄
    bt_cleantr.config(state='')#清除表格
    #'''
    #********** 讀取通訊參數初始值 *********
    port = port_E.get() #通信端口
    mode_E_a = mode_E.get() #通訊類型    
    baudrate = int(baudrate_E.get()) #傳輸速率
    bytesize = int(bytesize_E.get()) #字節大小
    parity = parity_E.get() #校驗
    stopbits = int(stopbits_E.get()) #停止位元
    timeout = float(timeout_E.get()) #超時

    bt_slaveaddress = int(bt_slaveaddress_E.get()) #豆溫從動裝置
    bt_PV_register = int(bt_register_PV_E.get()) #豆溫PV註冊位址
    bt_SV_register = int(bt_register_SV_E.get()) #豆溫SV註冊位址

    et_slaveaddress = int(et_slaveaddress_E.get()) #環境溫從動裝置
    et_PV_register = int(et_register_PV_E.get()) #環境溫PV註冊位址
    et_SV_register = int(et_register_SV_E.get()) #環境溫SV註冊位址

    entry_slaveaddress = int(entry_slaveaddress_E.get()) #入風溫從動裝置
    entry_PV_register = int(entry_register_PV_E.get()) #入風溫PV註冊位址
    entry_SV_register = int(entry_register_SV_E.get()) #入風溫SV註冊位址

    #********** 通訊參數初始值  END *********
    #'''
    #設定豆溫、環境溫之RoR以及程式執行圈數和座標SCALE的初始值
    counter_data = 0;counter_data_ax = 0;counter_data_temp = 0
    record_start_flag = 0 #開始記錄旗標
    charge_flag = 0  #入豆旗標
    fc_flag = 0 #一爆點旗標
    fc_start_time = 0   #一爆點起始時間
    rtp_flag = 0 #回溫點旗標
    dry_end_flag = 0 #脫水結束旗標
    gp_flag = 0 #金黃點旗標
    pre_base_time = 0 #估算脫水結束的基礎時間
    fc_end_flag = 0  #1爆結束旗標
    secondc_flag = 0 #2爆點旗標
    secondc_end_flag = 0 #2爆結束旗標.
    drop_flag = 0    #下豆點旗標
    ss_t0_flag = 0   #全息烘焙 T0 點
    ss_t1_flag = 0   #全息烘焙 T1 點
    ss_t2_flag = 0   #全息烘焙 T2 點
    agtron_temp = 0  #艾格狀數值基礎溫度
    agtron_num = 0  #艾格狀數值

    BT_ror = 0
    ET_ror = 0

    t0_temp = float(t0_temp_E_min.get())
    t1_temp = float(t1_temp_E_min.get())
    t2_temp = float(t2_temp_E_min.get())

    if (state_a == '0' or state_a == 'go') and charge_flag == 0: #防止按結束鍵時,將資料清空
        time_data = []; bt_temperature_data = []; et_temperature_data = []; entry_temperature_data = []
        event_data = []; ror_bt_data = []; ror_et_data= []
        step_data = []    #階段時間資料
        roast_time = ''
        ep_y_0 =0 ; ep_y_1 =[0,] ; ef_y_0 =0 ; ef_y_1 = [0,] 
        # round(float(slider_p.get()),0)  round(float(slider.get()),0)

    if et_slaveaddress == 0 and bt_slaveaddress == 0:   #模擬豆溫****************NOTE******
        BT = 190
        ET = 0.0
        entry = 0.0

    def catch_temp():
        #start = time.time() #估算程式執行一圈所需時間的起始時間值
        global time_data
        global bt_temperature_data
        global et_temperature_data
        global entry_temperature_data
        global event_data
        global ror_bt_data
        global ror_et_data
        global counter_data
        global ET
        global BT
        global entry
        global BT_ror
        global ET_ror
        global state_a
        global counter_data_ax
        global counter_data_temp
        global rt1  #入豆時間點
        global roast_time   #總烘豆時間
        global record_start_flag #開始記錄旗標
        global charge_flag  #入豆旗標
        global rtp_flag #回溫點旗標
        global dry_end_flag #脫水結束旗標
        global gp_flag  #金黃點旗標
        global fc_flag  #1爆點旗標
        global fc_end_flag  #1爆結束旗標
        global secondc_flag #2爆點旗標
        global secondc_end_flag #2爆結束旗標.
        global drop_flag    #下豆點旗標
        global pre_base_time #估算脫水結束的基礎時間
        global fc_start_time    #一爆點起始時間
        global step_data    #階段時間資料
        global ss_t0_flag   #全息烘焙 T0 點
        global ss_t1_flag   #全息烘焙 T1 點
        global ss_t2_flag   #全息烘焙 T2 點
        global agtron_num   #艾格狀數值
        global agtron_temp  #艾格狀數值基礎溫度
        global t0_temp  #T0點溫度
        global t1_temp  #T1點溫度
        global t2_temp  #T2點溫度
        global typess_no #烘焙方式選擇
        global ep_y_0
        global ep_y_1
        global ef_y_0
        global ef_y_1

        state_sc = ''
        sampling_time = 1 #取樣頻率(秒)
        #***** 設定溫度表頭參數 *****
        if et_slaveaddress != 0 or bt_slaveaddress != 0:#***** 設定溫度表頭參數
            #'''
            #***** 設定溫度表頭參數 *****  ==> 測試將*溫度表頭設定參數*放到另一模組(def)
            if mode_E_a == 'RTU' :
                mode= minimalmodbus.MODE_RTU
            elif mode_E_a == 'ASCII':
                mode= minimalmodbus.MODE_ASCII
            #Set up instrument
            instrument_bt = minimalmodbus.Instrument(port,bt_slaveaddress,mode)
            instrument_et = minimalmodbus.Instrument(port,et_slaveaddress,mode)
            instrument_entry = minimalmodbus.Instrument(port,entry_slaveaddress,mode)
    
            #Make the settings explicit
            instrument_bt.serial.baudrate = baudrate        # Baudrate
            instrument_bt.serial.bytesize = bytesize        # bytesize
            instrument_et.serial.baudrate = baudrate        # Baudrate
            instrument_et.serial.bytesize = bytesize        # bytesize
            instrument_entry.serial.baudrate = baudrate        # Baudrate
            instrument_entry.serial.bytesize = bytesize        # bytesize

            if parity == "EVEN":
                instrument_bt.serial.parity = minimalmodbus.serial.PARITY_EVEN   #parity
                instrument_et.serial.parity = minimalmodbus.serial.PARITY_EVEN   #parity
                instrument_entry.serial.parity = minimalmodbus.serial.PARITY_EVEN   #parity

            instrument_bt.serial.stopbits = stopbits        # stopbits
            instrument_bt.serial.timeout  = timeout         # timeout seconds
            instrument_et.serial.stopbits = stopbits        # stopbits
            instrument_et.serial.timeout  = timeout         # timeout seconds
            instrument_entry.serial.stopbits = stopbits        # stopbits
            instrument_entry.serial.timeout  = timeout         # timeout seconds

            # Good practice
            instrument_bt.close_port_after_each_call = True
            instrument_bt.clear_buffers_before_each_transaction = True
            instrument_et.close_port_after_each_call = True
            instrument_et.clear_buffers_before_each_transaction = True
            instrument_entry.close_port_after_each_call = True
            instrument_entry.clear_buffers_before_each_transaction = True

            #***** 設定溫度表頭參數  END *****
        #elif et_slaveaddress == 0 and bt_slaveaddress == 0:

        #讀取並顯示豆溫及環境溫之溫控器的SV值(表頭設定值==>可控制瓦斯On/Off用)
        #SV_temperature_bt = instrument_bt.read_register(bt_SV_register) /10    #豆溫表頭SV值
        #bt_temp_sv_l.config(text=str(SV_temperature_bt))                       #豆溫表頭SV值
        #SV_temperature_et = instrument_et.read_register(et_SV_register) /10    #環境溫表頭SV值
        #et_temp_sv_l.config(text=str(SV_temperature_et))                       #環境溫表頭SV值

        #顯示豆溫及環境溫(程式每執行一圈即顯示一次)  程式執行一回約0.1251秒
        if et_slaveaddress != 0 and bt_slaveaddress != 0 and entry_slaveaddress != 0:    #有豆溫、環境溫及入風溫
            BT =instrument_bt.read_register(bt_PV_register) / 10
            ET =instrument_et.read_register(et_PV_register) / 10
            entry =instrument_entry.read_register(entry_PV_register) / 10
        elif et_slaveaddress != 0 and bt_slaveaddress != 0:    #有豆溫及環境溫
            BT =instrument_bt.read_register(bt_PV_register) / 10
            ET =instrument_et.read_register(et_PV_register) / 10
            entry = 0.0
        elif et_slaveaddress == 0 and bt_slaveaddress != 0:   #只有豆溫
            BT =instrument_bt.read_register(bt_PV_register) / 10
            ET = 0.0
            entry = 0.0
        elif et_slaveaddress == 0 and bt_slaveaddress == 0:   #模擬豆溫****************NOTE******
            if counter_data_temp < 5 :
                BT = BT - 1.2
            elif counter_data_temp >= 5 and counter_data_temp < 90 :
                BT = BT - (90-counter_data_temp)/130
                ET = BT + 5
                entry = 200.0
            elif counter_data_temp >= 90 and counter_data_temp < 93 :
                BT = BT #- (counter_data_temp)/450
                ET = BT + 5
                entry = 205.0
            elif counter_data_temp >= 93 and counter_data_temp < 100 :
                BT = BT + (counter_data_temp)/500
                ET = BT + 5
                entry = 210.0
            elif counter_data_temp >= 100 and counter_data_temp < 150:
                BT = BT + (counter_data_temp/950)
                ET = BT + 5
                entry = 215.0
            elif counter_data_temp >= 150 and counter_data_temp < 360:
                BT = BT + (360-counter_data_temp)/850
                ET = BT + 5        
                entry = 220.0
            elif counter_data_temp >= 360 and counter_data_temp < 450:
                BT = BT + (450-counter_data_temp)/900
                ET = BT + 5
                entry = 225.0
            elif counter_data_temp >= 450 and counter_data_temp < 540:
                BT = BT + (540-counter_data_temp)/950
                ET = BT + 5
                entry = 230.0
            elif counter_data_temp >= 540 and counter_data_temp < 660:
                BT = BT + (660-counter_data_temp)/1000
                ET = BT + 5
                entry = 232.0

        BT = round(BT,1)
        ET = round(ET,1)
        entry = round(entry,1)
        bt_temp_l.config(text=str(BT))#顯示豆溫
        bt_temp_ss_l.config(text=str(BT))#顯示豆溫_ss
        et_temp_l.config(text=str(ET))#顯示環境溫
        entry_temp_l.config(text=str(entry))#顯示入風溫

        #************* RoR 計算 *************
        #10秒鐘過後，每秒計算1次RoR #取10秒平均計算一次
        #所以每10秒的溫升就是10/60分鐘的溫升
        if counter_data_temp >= 10  and counter_data_temp < 60 :
            BT_ror = round((BT - bt_temperature_data[counter_data_temp - 10])/(10/60),1)
            ET_ror = round((ET - et_temperature_data[counter_data_temp - 10])*6,1)
            bt_temp_ror_l.config(text=str(BT_ror))#顯示豆溫ROR
            bt_temp_ror_ss_l.config(text=str(BT_ror))#顯示豆溫ROR
            #et_temp_ror_l.config(text=str(ET_ror))#顯示環境溫ROR

        #60秒後，每秒計算1次RoR
        #由當下的溫度讀值減去第60秒前的讀值來做為ROR的數值
        elif counter_data_temp >= 60:
            BT_ror = round((BT - bt_temperature_data[counter_data-60]),1)
            ET_ror = round((ET - et_temperature_data[counter_data-60]),1)
            bt_temp_ror_l.config(text=str(BT_ror))
            bt_temp_ror_ss_l.config(text=str(BT_ror))#顯示豆溫ROR
            #et_temp_ror_l.config(text=str(ET_ror))
        #************* RoR 計算 End *************

        #****** 偵測按鍵事件並記錄 ******
        if state_a == '0' and charge_flag == 0:#開機狀態
            state_sc = ''
            state_a = 'open machine'
            bt_charge.config(state='')#入豆
            bt_records.config(state='')#開使記錄
            Radiobutton(frame4,text="北歐烘焙法",state=DISABLED).grid(row=0,column=0,columnspan=14,padx=5,pady=5)
            Radiobutton(frame4,text="Scott Rao_漸降式烘焙法",state=DISABLED).grid(row=0,column=15,columnspan=14,padx=5,pady=5)
            Radiobutton(frame4,text="小野善造_完全烘焙法",state=DISABLED).grid(row=15,column=0,columnspan=14,padx=5,pady=5)
            Radiobutton(frame4,text="全息烘焙法",state=DISABLED).grid(row=15,column=15,columnspan=14,padx=5,pady=5)

        elif state_a == 'go' and record_start_flag == 0 :#開始記錄
            bt_records.config(state=DISABLED)
            record_start_flag = 1 #開始記錄旗標

        elif (state_a == 'CHARGE' or ((counter_data_temp >= 1) and (bt_temperature_data[-1] - BT > 0.01))) and charge_flag == 0 and record_start_flag == 1:
            #
            #print(bt_temperature_data[-1] - BT)
            #將熱機時間的紀錄歸0
            if et_slaveaddress == 0 and bt_slaveaddress == 0:
                BT = 190 #模擬豆溫度 正式程式執行要取消掉*****************NOTE**********************
                ET = 190 #模擬環境溫度 正式程式執行要取消掉*****************NOTE********************
                entry = 200 #模擬入風溫度 正式程式執行要取消掉*****************NOTE*****************

            state_a = 'CHARGE'
            canvas.delete("all")
            draw_panal()
            #print(typess_no)
            draw_panal_ss(typess_no)
            counter_data = 0
            counter_data_temp = 0
            counter_data_ax = 0
            time_data = []
            bt_temperature_data = []
            et_temperature_data = []
            entry_temperature_data = []
            event_data=[]
            ror_bt_data=[]
            ror_et_data=[]
            rt1=time.time() #入豆時間點
            charge_flag = 1
            shinf = '   ' + str(BT)
            charge_E.insert(0,shinf)
            #--- 顯示入豆點 ---
            sp1_x = counter_data_temp+66
            sp1_y = 420-(BT*1.6)
            sp2_x = sp1_x + 20
            sp2_y = sp1_y - 20
            canvas.create_line(sp1_x,sp1_y,sp2_x,sp2_y,fill='blue')
            canvas.create_text(sp2_x,sp2_y,text='CHARGE',fill='red')#state=DISABLED,
            #Button(frame1,text='入豆',state=DISABLED,style='W.TButton', command=lambda:roast_state('CHARGE')).grid(row=8,column=0)
            bt_charge.config(state=DISABLED)
            #Button(frame1,text='回溫點',style='W2.TButton', command=lambda:roast_state('TP')).grid(row=9,column=0)
            bt_tp.config(state='')

        elif state_a =='FCs' and fc_flag == 0:
            fc_flag = 1
            fc_start_time = counter_data #First crack time start
            shinf = '   ' + str(BT) +' | '+ roast_time
            agtron_temp = BT #預估艾格狀數的基礎溫度
            fc_E.insert(0,shinf)
            step_data.append(counter_data_temp)    #梅納階段時間資料
            #--- 顯示一爆點 ---
            sp1_x = (counter_data_temp*0.5)+66
            sp1_y = 420-(BT*1.6)
            sp2_x = sp1_x + 20
            sp2_y = sp1_y + 40
            canvas.create_line(sp1_x,sp1_y,sp2_x,sp2_y,fill='blue')
            canvas.create_text(sp2_x,sp2_y,text='FCs',fill='red')
            bt_fcs.config(state=DISABLED)#一爆

            if typess_no == 1 :#估計北歐烘焙的下豆點範圍
                canvas_ss.create_rectangle(66+counter_data_temp/2+13, (420-190*1.6), 66+counter_data_temp/2+60, (420-210*1.6),fill='brown', stipple="gray50" )#下豆點區域

        elif state_a =='FCe' and fc_end_flag == 0:
            fc_end_flag = 1
            shinf = '   ' + str(BT) +' | '+ roast_time
            fcend_E.insert(0,shinf)
            #--- 顯示一爆結束點 ---
            sp1_x = (counter_data_temp*0.5)+66
            sp1_y = 420-(BT*1.6)
            sp2_x = sp1_x + 20
            sp2_y = sp1_y + 40
            canvas.create_line(sp1_x,sp1_y,sp2_x,sp2_y,fill='blue')
            canvas.create_text(sp2_x,sp2_y,text='FCe',fill='red')
            bt_fce.config(state=DISABLED)#一爆結束

        elif state_a =='SCs' and secondc_flag == 0:
            secondc_flag = 1
            shinf = '   ' + str(BT) +' | '+ roast_time
            sc_E.insert(0,shinf)
            #--- 顯示二爆點 ---
            sp1_x = (counter_data_temp*0.5)+66
            sp1_y = 420-(BT*1.6)
            sp2_x = sp1_x + 20
            sp2_y = sp1_y + 40
            canvas.create_line(sp1_x,sp1_y,sp2_x,sp2_y,fill='blue')
            canvas.create_text(sp2_x,sp2_y,text='SCs',fill='blue')
            bt_scs.config(state=DISABLED)#二爆

        elif state_a =='SCe' and secondc_end_flag == 0:
            secondc_end_flag = 1
            shinf = '   ' + str(BT) +' | '+ roast_time
            scend_E.insert(0,shinf)
            #--- 顯示二爆結束點 ---
            sp1_x = (counter_data_temp*0.5)+66
            sp1_y = 420-(BT*1.6)
            sp2_x = sp1_x - 40
            sp2_y = sp1_y - 40
            canvas.create_line(sp1_x,sp1_y,sp2_x,sp2_y,fill='blue')
            canvas.create_text(sp2_x,sp2_y,text='SCe',fill='red')
            bt_sce.config(state=DISABLED)#二爆結束

        elif ((counter_data_temp >= 5) and ((bt_temperature_data[-1] - BT > 1) or (state_a =='DROP'))) and (drop_flag == 0 and fc_flag == 1):
            drop_flag = 1
            bt_drop.config(state=DISABLED)#下豆
            shinf = '   ' + str(BT) +' | '+ roast_time
            drop_E.insert(0,shinf)
            step_data.append(counter_data_temp)    #發展階段時間資料
            #--- 顯示下豆點 ---
            sp1_x = (counter_data_temp*0.5)+66
            sp1_y = 420-(BT*1.6)
            sp2_x = sp1_x + 60
            sp2_y = sp1_y - 20
            canvas.create_line(sp1_x,sp1_y,sp2_x,sp2_y,fill='blue')
            canvas.create_text(sp2_x,sp2_y,text='DROP',fill='red')
        #****** 偵測按鍵事件並記錄 End ******

        #******* 自動事件紀錄 *******
        #print(len(bt_temperature_data))
        if ((len(bt_temperature_data) >= 2) and (charge_flag == 1) and (rtp_flag == 0)):
            if ((((bt_temperature_data[-2] >= bt_temperature_data[-1]) and (BT > bt_temperature_data[-1]))) or state_a == 'TP'):#
                state_a = 'TP'
                bt_tp.config(state=DISABLED)#回溫點
                rtp_flag = 1 #回溫點旗標
                shinf = '   ' + str(BT) +' | '+ roast_time
                rtp_E.insert(0,shinf)
                pre_base_time = counter_data_temp
                #--- 顯示回溫點 ---
                sp1_x = (counter_data_temp*0.5)+66
                sp1_y = 420-(BT*1.6)
                sp2_x = sp1_x - 20
                sp2_y = sp1_y + 20
                canvas.create_line(sp1_x,sp1_y,sp2_x,sp2_y,fill='blue')
                canvas.create_text(sp2_x,sp2_y,text='TP',fill='red')

                #*****----- 全息烘焙繪圖 回溫點 -----*****
                if typess_no == 4:
                    canvas_ss.create_rectangle((36 + counter_data_temp/2), (420-70*1.6), (126 + counter_data_temp/2), (420-90*1.6),fill='lightgreen', stipple="gray50" )#回溫點範圍
                    canvas_ss.create_line(36,420-(BT*1.6),(66 + counter_data_temp/2),420-(BT * 1.6),width=1,fill='#ff8000', dash=(10,2))
                    canvas_ss.create_line((66 + counter_data_temp/2),420-(BT*1.6),(66 + counter_data_temp/2),420,width=1,fill='#ff8000', dash=(10,2))
                    canvas_ss.create_text((10+36 + counter_data_temp/2),420-(BT * 1.6)+10,text='回溫點',fill='brown')
                #*****----- 全息烘焙繪圖 End -----*****
                bt_drye.config(state='')#脫水結束

        #---**** 預計脫水結束時間 ****---
        if rtp_flag == 1 and dry_end_flag == 0 and BT_ror > 0:
            pre_edr_time = pre_base_time+(counter_data_temp - pre_base_time) + ((150.0 - BT) / BT_ror * 60)
            pre_time_min = pre_edr_time // 60
            pre_time_sec = pre_edr_time % 60
            pre_dryend_time = str(pre_time_min) +'分'+str(round(pre_time_sec,1))+'秒'
            even_predryend_l.config(text = pre_dryend_time )

        if ((BT >= 150.0  or state_a == 'DRYe') and (dry_end_flag == 0) and rtp_flag == 1):
            state_a = 'DRYe'
            bt_drye.config(state=DISABLED)#脫水結束
            dry_end_flag = 1 #脫水結束旗標
            shinf = '   ' + str(BT) +' | '+ roast_time
            enddry_E.insert(0,shinf)
            step_data.append(counter_data_temp)    #脫水完成時間資料
            #--- 顯示脫水結束 ---
            sp1_x = (counter_data_temp*0.5)+66
            sp1_y = 420-(BT*1.6)
            sp2_x = sp1_x - 60
            sp2_y = sp1_y - 40
            canvas.create_line(sp1_x,sp1_y,sp2_x,sp2_y,fill='blue')
            canvas.create_text(sp2_x,sp2_y,text='DRYe',fill='blue')
            bt_gdp.config(state='')#金黃點

        if (BT >= 170.0  or state_a == 'GDp') and gp_flag == 0 and dry_end_flag == 1:
            state_a = 'GDp'
            gp_flag = 1 #金黃點旗標
            shinf = '   ' + str(BT) +' | '+ roast_time
            gp_E.insert(0,shinf)
            #--- 顯示金黃點 ---
            sp1_x = (counter_data_temp*0.5)+66
            sp1_y = 420-(BT*1.6)
            sp2_x = sp1_x - 40
            sp2_y = sp1_y - 40
            canvas.create_line(sp1_x,sp1_y,sp2_x,sp2_y,fill='blue')
            canvas.create_text(sp2_x,sp2_y,text='GDp',fill='blue')
            bt_gdp.config(state=DISABLED)#金黃點
        #---------- 全息烘焙繪圖 ----------
        if BT >= t0_temp and rtp_flag == 1 and ss_t0_flag == 0 and typess_no == 4:
            #*****----- 全息烘焙繪圖 T0點 -----*****
            ss_t0_flag = 1
            canvas_ss.create_rectangle((36 + counter_data_temp/2), (420-90*1.6), (126 + counter_data_temp/2), (420-100*1.6),fill='blue', stipple="gray50" )#T0點範圍
            canvas_ss.create_line(36,420-(BT*1.6),(66 + counter_data_temp/2),420-(BT * 1.6),width=1,fill='#ff8000', dash=(10,2))
            canvas_ss.create_line((66 + counter_data_temp/2),420-(BT*1.6),(66 + counter_data_temp/2),420,width=1,fill='#ff8000', dash=(10,2))
            canvas_ss.create_text((15+66 + counter_data_temp/2),420-(BT * 1.6)+10,text='T0點',fill='brown')
            #*****----- 全息烘焙繪圖 T0 End -----*****

        if BT >= t1_temp and rtp_flag == 1 and ss_t1_flag == 0 and typess_no == 4:
            #*****----- 全息烘焙繪圖 T1點 -----*****
            ss_t1_flag = 1
            canvas_ss.create_rectangle((36 + counter_data_temp/2), (420-110*1.6), (126 + counter_data_temp/2), (420-120*1.6),fill='#ff8000', stipple="gray50" )#T1點範圍
            canvas_ss.create_line(36,420-(BT*1.6),(66 + counter_data_temp/2),420-(BT * 1.6),width=1,fill='#ff8000', dash=(10,2))
            canvas_ss.create_line((66 + counter_data_temp/2),420-(BT*1.6),(66 + counter_data_temp/2),420,width=1,fill='#ff8000', dash=(10,2))
            canvas_ss.create_text((15+66 + counter_data_temp/2),420-(BT * 1.6)+10,text='T1點',fill='brown')
            #*****----- 全息烘焙繪圖 T1 End -----*****

        if BT >= t2_temp and rtp_flag == 1 and ss_t2_flag == 0 and typess_no == 4:
            #*****----- 全息烘焙繪圖 T2點 -----*****
            ss_t2_flag = 1
            canvas_ss.create_rectangle((36 + counter_data_temp/2), (420-170*1.6), (126 + counter_data_temp/2), (420-190*1.6),fill='red', stipple="gray50" )#T2點範圍
            canvas_ss.create_line(36,420-(BT*1.6),(66 + counter_data_temp/2),420-(BT * 1.6),width=1,fill='#ff8000', dash=(10,2))
            canvas_ss.create_line((66 + counter_data_temp/2),420-(BT*1.6),(66 + counter_data_temp/2),420,width=1,fill='#ff8000', dash=(10,2))
            canvas_ss.create_text((15+66 + counter_data_temp/2),420-(BT * 1.6)+10,text='T2點',fill='brown')
            #*****----- 全息烘焙繪圖 T2 End -----*****

            #*****----- 預估豆表艾格狀數值 -----*****
        if fc_flag == 1 and drop_flag == 0:
            agtron_num = float(fc_agtron_E.get())#預設一爆時的艾格狀數值
            #agtron_temp:預估艾格狀數的基礎溫度(一爆時的豆溫)
            #理論依據:一爆後溫度每上升1度，豆表色度加深2個焙度
            #所以當一爆時艾格狀數(agtron)-(當前豆溫-一爆時的豆溫)*2
            #所得到的數值，作為當下豆溫，豆表所呈現的色度
            pre_agtron = round((agtron_num - (BT - agtron_temp) * 2),1)
            agtron_L_S.config(text=str(pre_agtron))
            f1_agtron_L_S.config(text=str(pre_agtron))
            #*****----- 預估豆表艾格狀數值 End -----*****
        #---------- 全息烘焙繪圖 ----------

        #******* 自動事件紀錄 End *******

        #******* 壓差計及瓦斯壓力 *******
        if counter_data_temp > 1 :
            temp_state = state_a
            ep_y_0 = ep_y_1[-1]
            ef_y_0 = ef_y_1[-1]
            if temp_state[0] == 'P':
                ep_y_1.append(round(float(slider_p.get())*4,0))
            if temp_state[0] == 'F':
                ef_y_1.append(round(float(slider.get())*2,0))
        #******* 壓差計及瓦斯壓力 End *******

        state_sc = state_a #將事件轉成紀錄用變數
         
        #******* 顯示總烘焙時間及將各種烘豆相關數據分別存入各自陣列 *******
        counter_data = int(round(rtime(rt1),0)) #時間座標產生器 呼叫總時數程式 回傳烘豆總秒數
        if counter_data != counter_data_temp:   #因為程式執行一次僅需0.1251秒
            counter_data_temp = counter_data    #所以需確認每秒僅記錄一次資料

            counter_data_ax = counter_data_temp*(0.5) #產生座標軸實際分鐘數 / 程式執行一回約0.0016秒
            #與總烘豆時間同步計時 ---此處時間用來做及時繪圖的數據
            time_min = counter_data_temp // 60
            time_sec = counter_data_temp % 60
            roast_time = str(time_min) +':'+str(time_sec)+''
            #print(roast_time)
            #將程式執行一次的各種烘豆相關數據分別存入各自陣列
            time_data.append(counter_data_ax)   #時間資料
            bt_temperature_data.append(BT)      #豆溫資料
            et_temperature_data.append(ET)      #環境溫資料
            entry_temperature_data.append(entry)      #入風溫資料

            ror_bt_data.append(BT_ror)          #豆溫ror資料
            ror_et_data.append(ET_ror)          #環境溫ror資料
            event_data.append(state_sc)         #事件資料

            #********** 將資料寫入 TreeView Table 中 **********
            tree.insert('',index=0,text=str(counter_data_temp),values=(str(counter_data_temp),str(BT),str(BT_ror),str(ET),str(ET_ror),roast_time,state_sc))
        
            state_a = '0'    #將狀態回復到開機

        #**********--- 繪圖區段 ---***********
        if counter_data_temp > 2:   #繪製溫度曲線
            time_data_1 = time_data[counter_data_temp-2]
            time_data_2 = time_data[counter_data_temp-1]
            bt_temperature_data_1 = bt_temperature_data[-2]*1.6     #乘以1.6是因為要符合座標軸比例關係
            bt_temperature_data_2 = bt_temperature_data[-1]*1.6
            et_temperature_data_1 = et_temperature_data[-2]*1.6
            et_temperature_data_2 = et_temperature_data[-1]*1.6
            entry_temperature_data_1 = entry_temperature_data[-2]*1.6
            entry_temperature_data_2 = entry_temperature_data[-1]*1.6

            canvas.create_line(66+time_data_1,420-bt_temperature_data_1,66+time_data_2,420-bt_temperature_data_2,width=2,fill='red')
            canvas.create_line(66+time_data_1,420-et_temperature_data_1,66+time_data_2,420-et_temperature_data_2,width=2,fill='blue')
            canvas.create_line(66+time_data_1,420-entry_temperature_data_1,66+time_data_2,420-entry_temperature_data_2,width=2,fill='green')

            canvas_ss.create_line(66+time_data_1,420-bt_temperature_data_1,66+time_data_2,420-bt_temperature_data_2,width=2,fill='red')
            canvas_ss.create_line(66+time_data_1,420-et_temperature_data_1,66+time_data_2,420-et_temperature_data_2,width=2,fill='blue')
            canvas_ss.create_line(66+time_data_1,420-entry_temperature_data_1,66+time_data_2,420-entry_temperature_data_2,width=2,fill='green')
            
            #繪出壓差及瓦斯壓力  
            canvas_event.create_line(66+time_data_1,110-ep_y_0,66+time_data_2,110-ep_y_1[-1],width=2,fill='red')
            canvas_event.create_line(66+time_data_1,110-ef_y_0,66+time_data_2,110-ef_y_1[-1],width=2,fill='blue')

        if counter_data_temp >= 10 :    #10秒鐘過後，繪製RoR曲線
            ror_bt_1 = ror_bt_data[-2] * 10*0.8 #乘以0.8是因為要符合座標軸比例關係
            ror_bt_2 = ror_bt_data[-1] * 10*0.8
            ror_et_1 = ror_et_data[-2] * 10*0.8
            ror_et_2 = ror_et_data[-1] * 10*0.8
            if ror_bt_2 >= 0 :
                canvas.create_line(66+time_data_1,420-ror_bt_1,66+time_data_2,420-ror_bt_2,width=2,fill='green')
                canvas_ss.create_line(66+time_data_1,420-ror_bt_1,66+time_data_2,420-ror_bt_2,width=2,fill='green')

            #if ror_et_2 >= 0 :
            #    canvas.create_line(66+time_data_1,420-ror_et_1,66+time_data_2,420-ror_et_2,width=2,fill='blue')
            #    canvas_ss.create_line(66+time_data_1,420-ror_et_1,66+time_data_2,420-ror_et_2,width=2,fill='blue')

        #*********** 繪圖區段 END ***********

        #*********** 計算並顯示一爆後發展時間 ***********
        if fc_flag == 1 and drop_flag == 0:
           fc_total_sec = counter_data_temp - fc_start_time
           fc_time_min = fc_total_sec // 60
           fc_time_sec = fc_total_sec % 60
           fc_roast_time = str(int(fc_time_min)) +':'+str(fc_time_sec)
           dev_p = str(float(round((100*fc_total_sec/counter_data_temp),2))) + ' %'
           even_dev_l.config(text=fc_roast_time) #顯示FC發展烘焙時間
           even_dev_p_l.config(text=dev_p)

        state_sc = ''   #將事件狀態變數清空
        greenbean_inf_L.after((sampling_time*300),catch_temp) #1000 = 1 秒 #跳脫繪圖狀態的隱藏開關
        
        #*********** 按下豆才繪出個階段占比長條圖 ***********
        if drop_flag == 1 : 
            #繪出個階段占比長條圖 / 顯示各階段所佔時間百分比
            p_s = 66
            p_d = 66 + step_data[0]*0.5
            p_m = 66 + step_data[1]*0.5
            p_f = 66 + step_data[2]*0.5
            pp_d = round((step_data[0] / step_data[2] * 100),2)
            pp_m = round((step_data[1] - step_data[0]) / step_data[2] * 100,2)
            pp_f = round((step_data[2] - step_data[1]) / step_data[2] * 100,2)
            pp_d_m = step_data[0] // 60
            pp_d_s = step_data[0] % 60
            pp_m_m = (step_data[1] - step_data[0]) // 60
            pp_m_s = (step_data[1] - step_data[0]) % 60
            pp_f_m = (step_data[2] - step_data[1]) // 60
            pp_f_s = (step_data[2] - step_data[1]) % 60

            #繪出個階段占比長條圖
            canvas.create_line(p_s, 20, p_d, 20, width=8, fill='green')  #脫水時間
            canvas.create_line(p_d, 20, p_m, 20, width=8, fill='#FF8800')    #梅納時間
            canvas.create_line(p_m, 20, p_f, 20, width=8, fill='brown')  #發展時間

            canvas_ss.create_line(p_s, 20, p_d, 20, width=8, fill='green')  #脫水時間
            canvas_ss.create_line(p_d, 20, p_m, 20, width=8, fill='#FF8800')    #梅納時間
            canvas_ss.create_line(p_m, 20, p_f, 20, width=8, fill='brown')  #發展時間

            #顯示各階段所佔時間百分比
            canvas.create_text(p_s + 40,40,text=str(pp_d)+' %',fill='green')
            canvas.create_text(p_d + 40,40,text=str(pp_m)+' %',fill='#FF8800')
            canvas.create_text(p_m + 40,40,text=str(pp_f)+' %',fill='brown')

            canvas_ss.create_text(p_s + 40,40,text=str(pp_d)+' %',fill='green')
            canvas_ss.create_text(p_d + 40,40,text=str(pp_m)+' %',fill='#FF8800')
            canvas_ss.create_text(p_m + 40,40,text=str(pp_f)+' %',fill='brown')

            canvas.create_text(p_s + 40,10,text=str(pp_d_m) + ' 分' + str(pp_d_s) + ' 秒',fill='green')
            canvas.create_text(p_d + 40,10,text=str(pp_m_m) + ' 分' + str(pp_m_s) + ' 秒',fill='#FF8800')
            canvas.create_text(p_m + 40,10,text=str(pp_f_m) + ' 分' + str(pp_f_s) + ' 秒',fill='brown')

            canvas_ss.create_text(p_s + 40,10,text=str(pp_d_m) + ' 分' + str(pp_d_s) + ' 秒',fill='green')
            canvas_ss.create_text(p_d + 40,10,text=str(pp_m_m) + ' 分' + str(pp_m_s) + ' 秒',fill='#FF8800')
            canvas_ss.create_text(p_m + 40,10,text=str(pp_f_m) + ' 分' + str(pp_f_s) + ' 秒',fill='brown')

        #end = time.time() #估算程式執行一圈所需時間的終止時間值
        #print(str(end - start)) #計算程式執行一圈所需時間
     
    catch_temp()
 
    if state_arg == 1 : #結束紀錄 
        greenbean_inf_L.after_cancel(greenbean_inf_L)
        greenbean_inf_L.after(0,lambda:greenbean_inf_L.destroy())
        msg_save = messagebox.askyesnocancel('Messagebox','是否存檔或取消?')
        if msg_save == True :
            save_data(time_data,bt_temperature_data,ror_bt_data,et_temperature_data,ror_et_data,entry_temperature_data,event_data,step_data)
            bt_recorde.config(state=DISABLED)
        elif msg_save == False :
            bt_recorde.config(state=DISABLED)
    return 1

def argument_setup(*args) -> None: #通訊參數設定儲存
    if args[0] == 0 : #由frame3寫入frame2
        port_E.delete(0,END)   #通信端口
        port_E.insert(0,port_cb.get())
        mode_E.delete(0,END) #通訊類型
        mode_E.insert(0,mode_cb.get())
        baudrate_E.delete(0,END) #傳輸速率
        baudrate_E.insert(0,baudrate_cb.get())
        bytesize_E.delete(0,END) #字節大小
        bytesize_E.insert(0,bytesize_cb.get())
        parity_E.delete(0,END) #校驗
        parity_E.insert(0,parity_cb.get())
        stopbits_E.delete(0,END) #停止位元
        stopbits_E.insert(0,stopbits_cb.get())
        timeout_E.delete(0,END) #超時
        timeout_E.insert(0,timeout_cb.get())
        bt_slaveaddress_E.delete(0,END) #豆溫從動裝置
        bt_slaveaddress_E.insert(0,bt_slaveaddress_cb.get())
        bt_register_PV_E.delete(0,END) #PV註冊位址
        bt_register_PV_E.insert(0,bt_register_PV_cb.get())
        bt_register_SV_E.delete(0,END) #SV註冊位址
        bt_register_SV_E.insert(0,bt_register_SV_cb.get())
        et_slaveaddress_E.delete(0,END) #豆溫從動裝置
        et_slaveaddress_E.insert(0,et_slaveaddress_cb.get())
        et_register_PV_E.delete(0,END) #PV註冊位址
        et_register_PV_E.insert(0,et_register_PV_cb.get())
        et_register_SV_E.delete(0,END) #SV註冊位址
        et_register_SV_E.insert(0,et_register_SV_cb.get())
        entry_slaveaddress_E.delete(0,END) #入風溫從動裝置
        entry_slaveaddress_E.insert(0,entry_slaveaddress_cb.get())
        entry_register_PV_E.delete(0,END) #PV註冊位址
        entry_register_PV_E.insert(0,entry_register_PV_cb.get())
        entry_register_SV_E.delete(0,END) #SV註冊位址
        entry_register_SV_E.insert(0,entry_register_SV_cb.get())

        notebook.select(1)
    elif args[0] == 1:#由frame3寫入檔案
        filename = equipment_name.get()
        comm_args = []
        comm_args.append(port_cb.get())#通信端口
        comm_args.append(mode_cb.get())#通訊類型
        comm_args.append(baudrate_cb.get())#傳輸速率
        comm_args.append(bytesize_cb.get())#字節大小
        comm_args.append(parity_cb.get())#校驗
        comm_args.append(stopbits_cb.get())#停止位元
        comm_args.append(timeout_cb.get())#超時
        comm_args.append(bt_slaveaddress_cb.get())#豆溫從動裝置
        comm_args.append(bt_register_PV_cb.get())#PV註冊位址
        comm_args.append(bt_register_SV_cb.get())#SV註冊位址
        comm_args.append(et_slaveaddress_cb.get())#環境溫從動裝置
        comm_args.append(et_register_PV_cb.get())#PV註冊位址
        comm_args.append(et_register_SV_cb.get())#SV註冊位址
        comm_args.append(entry_slaveaddress_cb.get())#入風溫從動裝置
        comm_args.append(entry_register_PV_cb.get())#PV註冊位址
        comm_args.append(entry_register_SV_cb.get())#SV註冊位址

        with open( filename +'.arg','w') as f:
            for arg in comm_args:
                f.write(arg)
                f.write('\n')
        #print(comm_args)
        messagebox.showinfo('information', '存檔完成')
    elif args[0] == 3:#由檔案寫入frame2
        argument_load =[]
        s_f=args[1].get()
        select_file_E.delete(0,END)
        select_file_E.insert(0,s_f)
        with open( s_f,'r') as f: 
            for arg_load in f.readlines() :
                argument_load.append(arg_load)
        #print(argument_load[0][:-1])
        port_E.delete(0,END)   #通信端口
        port_E.insert(0,argument_load[0][:-1])
        mode_E.delete(0,END) #通訊類型
        mode_E.insert(0,argument_load[1][:-1])
        baudrate_E.delete(0,END) #傳輸速率
        baudrate_E.insert(0,argument_load[2][:-1])
        bytesize_E.delete(0,END) #字節大小
        bytesize_E.insert(0,argument_load[3][:-1])
        parity_E.delete(0,END) #校驗
        parity_E.insert(0,argument_load[4][:-1])
        stopbits_E.delete(0,END) #停止位元
        stopbits_E.insert(0,argument_load[5][:-1])
        timeout_E.delete(0,END) #超時
        timeout_E.insert(0,argument_load[6][:-1])
        bt_slaveaddress_E.delete(0,END) #豆溫從動裝置
        bt_slaveaddress_E.insert(0,argument_load[7][:-1])
        bt_register_PV_E.delete(0,END) #PV註冊位址
        bt_register_PV_E.insert(0,argument_load[8][:-1])
        bt_register_SV_E.delete(0,END) #SV註冊位址
        bt_register_SV_E.insert(0,argument_load[9][:-1])
        et_slaveaddress_E.delete(0,END) #環境溫從動裝置
        et_slaveaddress_E.insert(0,argument_load[10][:-1])
        et_register_PV_E.delete(0,END) #PV註冊位址
        et_register_PV_E.insert(0,argument_load[11][:-1])
        et_register_SV_E.delete(0,END) #SV註冊位址
        et_register_SV_E.insert(0,argument_load[12][:-1])
        entry_slaveaddress_E.delete(0,END) #入風溫從動裝置
        entry_slaveaddress_E.insert(0,argument_load[13][:-1])
        entry_register_PV_E.delete(0,END) #PV註冊位址
        entry_register_PV_E.insert(0,argument_load[14][:-1])
        entry_register_SV_E.delete(0,END) #SV註冊位址
        entry_register_SV_E.insert(0,argument_load[15][:-1])
       
        notebook.select(1)
    elif args[0] == 4:#由檔案讀入已存在烘豆資料
        roast_data_load =[]
        r_f=args[1].get()
        with open( r_f,'r') as f: 
            for roast_load in f.readlines() :
                roast_data_load.append(roast_load)
        if args[2] == 1 :
            redraw_profile(roast_data_load,r_f) #呼叫歷史檔案曲線繪圖模組
        elif args[2] == 2 : #呼叫轉成烘焙紀錄表模組
            record_table(roast_data_load,r_f)
    elif args[0] == 5:#測試連線參數是否正確可以讀到控制器的數值
        try:
            #********** 讀取通訊參數初始值 *********
            port = port_E.get() #通信端口
            mode_E_a = mode_E.get() #通訊類型    
            baudrate = int(baudrate_E.get()) #傳輸速率
            bytesize = int(bytesize_E.get()) #字節大小
            parity = parity_E.get() #校驗
            stopbits = int(stopbits_E.get()) #停止位元
            timeout = float(timeout_E.get()) #超時

            bt_slaveaddress = int(bt_slaveaddress_E.get()) #豆溫從動裝置
            bt_PV_register = int(bt_register_PV_E.get()) #豆溫PV註冊位址
            bt_SV_register = int(bt_register_SV_E.get()) #豆溫SV註冊位址

            et_slaveaddress = int(et_slaveaddress_E.get()) #環境溫從動裝置
            et_PV_register = int(et_register_PV_E.get()) #環境溫PV註冊位址
            et_SV_register = int(et_register_SV_E.get()) #環境溫SV註冊位址

            entry_slaveaddress = int(et_slaveaddress_E.get()) #入風溫從動裝置
            entry_PV_register = int(et_register_PV_E.get()) #入風溫PV註冊位址
            entry_SV_register = int(et_register_SV_E.get()) #入風溫SV註冊位址

            #print(bt_slaveaddress, '   ' ,et_slaveaddress)
            #**********  通訊參數初始值  END *********
            #***** 溫度表頭設定參數 *****
            if (et_slaveaddress != 0 or bt_slaveaddress != 0):
                if mode_E_a == 'RTU' :
                    mode= minimalmodbus.MODE_RTU
                elif mode_E_a == 'ASCII':
                    mode= minimalmodbus.MODE_ASCII
                #Set up instrument
                instrument_bt = minimalmodbus.Instrument(port,bt_slaveaddress,mode)
                instrument_et = minimalmodbus.Instrument(port,et_slaveaddress,mode)
                instrument_entry = minimalmodbus.Instrument(port,entry_slaveaddress,mode)
    
                #Make the settings explicit
                instrument_bt.serial.baudrate = baudrate        # Baudrate
                instrument_bt.serial.bytesize = bytesize        # bytesize
                instrument_et.serial.baudrate = baudrate        # Baudrate
                instrument_et.serial.bytesize = bytesize        # bytesize
                instrument_entry.serial.baudrate = baudrate        # Baudrate
                instrument_entry.serial.bytesize = bytesize        # bytesize

                if parity == "EVEN":
                    instrument_bt.serial.parity = minimalmodbus.serial.PARITY_EVEN   #parity
                    instrument_et.serial.parity = minimalmodbus.serial.PARITY_EVEN   #parity
                    instrument_entry.serial.parity = minimalmodbus.serial.PARITY_EVEN   #parity

                instrument_bt.serial.stopbits = stopbits        # stopbits
                instrument_bt.serial.timeout  = timeout         # timeout seconds
                instrument_et.serial.stopbits = stopbits        # stopbits
                instrument_et.serial.timeout  = timeout         # timeout seconds
                instrument_entry.serial.stopbits = stopbits        # stopbits
                instrument_entry.serial.timeout  = timeout         # timeout seconds

                # Good practice
                instrument_bt.close_port_after_each_call = True
                instrument_bt.clear_buffers_before_each_transaction = True
                instrument_et.close_port_after_each_call = True
                instrument_et.clear_buffers_before_each_transaction = True
                instrument_entry.close_port_after_each_call = True
                instrument_entry.clear_buffers_before_each_transaction = True

            #*****溫度表頭設定參數  END*****
            #讀取並顯示豆溫及環境溫之溫控器的PV值

            if (et_slaveaddress == 0 and bt_slaveaddress != 0):
                TEST_BT =instrument_bt.read_register(bt_PV_register) / 10
                TEST_BT = round(TEST_BT,1)
                msgtxt = 'OK  豆溫: '+str(TEST_BT)
                messagebox.showinfo('測試結果', msgtxt)
                #Button(frame1,text="開機", style='W.TButton',command=lambda:temp_ror('0')).grid(row=2,column=0,columnspan=2,padx=5,pady=5)
                bt_openmc.config(state='')
            elif (et_slaveaddress != 0 and bt_slaveaddress != 0):
                TEST_BT =instrument_bt.read_register(bt_PV_register) / 10
                TEST_ET =instrument_et.read_register(et_PV_register) / 10
                TEST_BT = round(TEST_BT,1)
                TEST_ET = round(TEST_ET,1)
                msgtxt = 'OK  豆溫: '+str(TEST_BT)+'   環境溫: '+str(TEST_ET)
                messagebox.showinfo('測試結果', msgtxt)
                #Button(frame1,text="開機", style='W.TButton',command=lambda:temp_ror('0')).grid(row=2,column=0,columnspan=2,padx=5,pady=5)    
                bt_openmc.config(state='')
            elif (et_slaveaddress != 0 and bt_slaveaddress != 0 and entry_slaveaddress != 0):
                TEST_BT =instrument_bt.read_register(bt_PV_register) / 10
                TEST_ET =instrument_et.read_register(et_PV_register) / 10
                TEST_entry =instrument_entry.read_register(entry_PV_register) / 10
                TEST_BT = round(TEST_BT,1)
                TEST_ET = round(TEST_ET,1)
                TEST_entry = round(TEST_entry,1)

                msgtxt = 'OK  豆溫: '+str(TEST_BT)+'   環境溫: '+str(TEST_ET)+'   入風溫: '+str(TEST_entry)
                messagebox.showinfo('測試結果', msgtxt)
                #Button(frame1,text="開機", style='W.TButton',command=lambda:temp_ror('0')).grid(row=2,column=0,columnspan=2,padx=5,pady=5)    
                bt_openmc.config(state='')
            elif (et_slaveaddress != 0 and bt_slaveaddress == 0):
                #TEST_BT =instrument_bt.read_register(bt_PV_register) / 10
                TEST_ET =instrument_et.read_register(et_PV_register) / 10
                #TEST_BT = round(TEST_BT,1)
                TEST_ET = round(TEST_ET,1)
                msgtxt = 'Beware!  豆溫:設定有問題!!    環境溫: ' + str(TEST_ET)
                messagebox.showinfo('測試結果', msgtxt)
                #Button(frame1,text="開機", style='W.TButton',command=lambda:temp_ror('0')).grid(row=2,column=0,columnspan=2,padx=5,pady=5)    
                bt_openmc.config(state='')
            elif (et_slaveaddress == 0 and bt_slaveaddress == 0):
                msgtxt = '豆溫: 模擬溫度'
                messagebox.showinfo('測試結果', msgtxt)
                #Button(frame1,text="開機", style='W.TButton',command=lambda:temp_ror('0')).grid(row=2,column=0,columnspan=2,padx=5,pady=5)    
                bt_openmc.config(state='')
        except: #Exception
            messagebox.showinfo('通訊參數測試訊息', '不知道怎麼了，反正發生錯誤惹')
    elif args[0] == 6:#讀取Artisan檔案
        r_f_as = args[1].get() #讀取Artisan檔名
        filename_ar = r_f_as[:-5]+".rxt" #設定轉檔後的檔名
        with open( r_f_as,'r') as f: 
            artisan_load = f.readlines()
        print(artisan_load)
        
        r=1 ; artisan_data_load = [] ; temp_array_event = []
        temp_array_timex = [] ; temp_array_temp1 = [] ; temp_array_temp2 = []
        ttt = ''; zzz = []
        for k in range(len(artisan_load[0])) :
            if artisan_load[0][k] == ":" :
                temp_at1 = (artisan_load[0][r:k+1]).split("'")
                #print(temp_at1) #觀察原始檔內容
                artisan_data_load += temp_at1
                r = k+1
        artisan_data_load += artisan_load[0][r:-1].split("'")
        status_type = {'CHARGE_BT','TP_time','TP_BT','DRY_time','DRY_BT','FCs_time','FCs_BT','FCe_time','FCe_BT',\
                        'SCs_time','SCs_BT','SCe_time','SCe_BT','DROP_time','DROP_BT','totaltime','dryphasetime',\
                        'midphasetime','finishphasetime','roastisodate','beans'}
        status_type_2 = {'timex','temp1','temp2'}#{時間軸,ET,BT}

        for item_s in status_type :#讀取{時間軸,ET,BT}以外的值
            if item_s == 'roastisodate' or item_s == 'beans':#烘豆日期與生豆名稱的格式較特殊
                add_index = 3
            else :
                add_index = 2
            try:#防止原始資料中沒有記錄到資料 Ex:FCe,SCs,SCe ... 等等。
                item_index = artisan_data_load.index(item_s)
                bean_name = list(artisan_data_load[item_index+add_index].split(","))
                if item_s == 'beans':
                    fff = (bean_name[0].strip()).split('\\')
                    ff1 = ''
                    for ttt in (fff):
                        if ttt != '':
                            ff1=ff1+'\\'+ttt
                    temp_array_event.append(item_s)
                    temp_array_event.append(ff1)

                    '''
                    #----- unicode 列印測試 -----
                    s=['\u5df4\u897f \u559c\u62c9\u6735']
                    s1 =['\u8863\u7d22\u6bd4\u4e9e \u53e4\u5409 \u7f55\u8c9d\u62c9 \u73ab\u7470\u8318\u8318']
                    str=s[0]#.decode('unicode_utf-8')  #.encode("EUC_KR")
                    str1=s1[0]
                    print(str,str1)
                    #----- unicode 列印測試 End -----
                    '''
                else :
                    temp_array_event.append(item_s)
                    temp_array_event.append(bean_name[0].strip())
                    #print(item_s, ' ', bean_name[0])
            except:
                temp_array_event.append(item_s)
                temp_array_event.append('')
                #print('no ',item_s)


        for item_s in status_type_2 :#讀取{時間軸,ET,BT}
            www = []
            #print(item_s)
            yyy = artisan_data_load.index(item_s)
            zzz = list(artisan_data_load[yyy+2].split(","))
            for j in range(len(zzz)-1):
                ttt = ((zzz[j].strip()).strip('[')).strip(']')
                if item_s == 'timex':
                    ttt = round(float(ttt),)
                else:
                    ttt = round(float(ttt),1)
                www.append(ttt)
            
            del zzz

            if item_s == 'temp2':
                temp_array_temp2 = www.copy()
            elif item_s == 'temp1':
                temp_array_temp1 = www.copy()
            elif item_s == 'timex':
                temp_array_timex = www.copy()

        print('temp1',temp_array_temp1)
        print('temp2',temp_array_temp2)
        print('timex',temp_array_timex)
        print('event',temp_array_event)
        print(temp_array_event[temp_array_event.index('beans')+1])

        #rxt檔案的存檔順序:
        #1.烘焙日期 2.產品名稱  3.生豆產地  4.生豆名稱  5.水分含量 
        #6.生豆密度 7.批次重量  8.設備名稱  9.設備容量  10.使用能源
        #11.操作人員    12.天氣溫度
        #13.序號:豆溫:豆溫ror:環境溫:環境溫ror:入風溫:時間:事件
            
        with open(filename_ar,'w') as f_a:
            f_a.write(temp_array_event[temp_array_event.index('roastisodate')+1])  #烘焙日期
            f_a.write('\n')
            f_a.write(temp_array_event[temp_array_event.index('beans')+1])   #產品名稱
            f_a.write('\n')
            f_a.write('')   #生豆產地
            f_a.write('\n')
            f_a.write('')   #生豆名稱
            f_a.write('\n')
            f_a.write('')   #水分含量
            f_a.write('\n')
            f_a.write('')   #生豆密度
            f_a.write('\n')
            f_a.write('')   #批次重量
            f_a.write('\n')
            f_a.write('')   #設備名稱
            f_a.write('\n')
            f_a.write('')   #設備容量
            f_a.write('\n')
            f_a.write('')   #使用能源
            f_a.write('\n')
            f_a.write('')   #操作人員
            f_a.write('\n')
            f_a.write('')   #天氣溫度
            f_a.write('\n')

            #存檔順序 序號 :豆溫:豆溫ror:環境溫:環境溫ror:入風溫:時間:事件
            for i in range(len(temp_array_timex)):   
                f_a.write(str(i+1)+':') #序號
                f_a.write(str(temp_array_temp2[i])) #豆溫
                f_a.write(':')
                f_a.write(str('0')) #豆溫ror
                f_a.write(':')
                f_a.write(str(temp_array_temp1[i])) #環境溫
                f_a.write(':')
                f_a.write(str('0')) #環境溫ror
                f_a.write(':')
                f_a.write(str('0')) #入風溫
                f_a.write(':')
                f_a.write(str(round((temp_array_timex[i] / 0.5),1)))    #時間
                f_a.write(':')
                f_a.write(str('0')) #事件
                f_a.write('\n')
            try:
                f_a.write(str(temp_array_event[temp_array_event.index('dryphasetime')+1]))
                f_a.write('\n')
                f_a.write(str(temp_array_event[temp_array_event.index('midphasetime')+1]))
                f_a.write('\n')
                f_a.write(str(temp_array_event[temp_array_event.index('finishphasetime')+1]))
                f_a.write('\n')
            except:
                messagebox.showinfo('alog to rxt', '資料不齊全，無法計算階段百分比')
        messagebox.showinfo('alon to rxt', '檔案轉換完成')
    return

def sl_f_ch(source):#壓差計選擇設定
    #print(type(float(source)))
    sl_value = round(float(source),0)
    f_value = 'F'+str(sl_value)
    slf_ch_b=Button(frame1,text=f_value,style='Wf.TButton',command=lambda:roast_state(f_value)).grid(row=18,column=5)
    roast_state(f_value)
    return 
    
def sl_p_ch(source):#瓦斯壓力選擇設定load_file_data
    #print(type(float(source)))
    sl_value = round(float(source),0)
    p_value = 'P'+str(sl_value)
    slp_ch_b=Button(frame1,text=p_value,style='Wf.TButton', command=lambda:roast_state(p_value)).grid(row=19,column=5)
    roast_state(p_value)
    #print(slider_p.get())
    return

def redraw_profile(roast_data_load,r_f):#歷史檔案曲線繪圖
    root_redraw_profile = Tk()
    root_redraw_profile.title(r_f)
    screenwidth = root_redraw_profile.winfo_screenwidth()
    screenheight = root_redraw_profile.winfo_screenheight()
    w_win = 1500
    h_win = 780
    x_offset = (screenwidth - w_win) / 2
    y_offset = ((screenheight - h_win) / 2)# - 30
    root_redraw_profile.title("Shokunin Coffee Roaster")
    root_redraw_profile.geometry("%dx%d+%d-%d" %(w_win,h_win,x_offset,y_offset))#1520
    root_redraw_profile.configure(bg='lightgreen')

    ttk.Style().configure("Line.TSeparator", background="#ff0000")
    canvas =Canvas(root_redraw_profile,width=1480,height=660,bg='#FEFEFE')#white
    canvas.grid(row=0,rowspan=11 ,column=0,columnspan=15,padx=5,pady=5)

    #第一座標Y軸
    canvas.create_line(40,0,40,640,width=2,fill='blue')#Y 第一座標軸
    for i in range(100,600,100):
        canvas.create_line(40,i,50,i,width=2,fill='red')
        canvas.create_text(20,i,text=str((6-(i // 100))*50),fill='blue')
        if i > 20 :
            canvas.create_line(51,i,1420,i,width=2,fill='#0f1', dash=(10,2))

    #第二座標Y軸
    canvas.create_line(1420,0,1420,640,width=2,fill='red')#Y 第二座標軸
    for j in range(100,600,100):
        canvas.create_line(1410,j,1420,j,width=2,fill='red')
        canvas.create_text(1440,j,text=str((6-(j // 100))*10),fill='red')

    #X座標軸
    canvas.create_line(40,600,1420,600,width=2,fill='green')#X 時間軸
    jup = 0    
    for k in range(100,1480,60):
        jup += 1
        if jup == 5 or jup == 10 or jup == 15 or jup == 20 or jup == 25 :
            canvas.create_line(k,600,k,590,width=2,fill='red')
            canvas.create_text(k,620,text=str(jup),fill='blue')
        else:
            canvas.create_line(k,600,k,590,width=2,fill='green') 

    #階段
    dry_start = float(rostep_dry_start_E.get())#內定脫水起始溫度 
    dry_end = float(rostep_dry_end_E.get())#內定脫水結束溫度
    maillard_start = float(rostep_maillard_start_E.get())#內定梅納階段起始溫度
    maillard_end = float(rostep_maillard_end_E.get())#內定梅納階段結束溫度
    development_start = float(rostep_development_start_E.get())#內定完成起始溫度
    development_end = float(rostep_development_end_E.get())#內定完成結束溫度

    canvas.create_rectangle(40, (600-dry_start*2), 1420, (600-dry_end*2),fill='lightgreen', stipple="gray25" )#脫水期activefill='lightgreen',activestipple="gray50"
    canvas.create_rectangle(40, (600-maillard_start*2), 1420, (600-maillard_end*2),fill='#ff8000', stipple="gray25" )#梅納期skyblue
    canvas.create_rectangle(40, (600-development_start*2), 1420, (600-development_end*2),fill='gray', stipple="gray25" )#發展期skyblue

    #**********繪圖區段***********
    time_data = []
    bt_temperature_data =[]
    ror_bt = []
    et_temperature_data =[]
    ror_et = []
    entry_temperature_data =[]
    event_data = []
    step_data = []    
    for i in range(17,len(roast_data_load)-3):
        reda = roast_data_load[i].split(':')
        #print(reda)
        bt_temperature_data.append(float(reda[1])*2)
        ror_bt.append(0 if (float(reda[2])*10*1) <= 0 else (float(reda[2])*10*1))
        et_temperature_data.append(float(reda[3])*2)
        ror_et.append(0 if (float(reda[4])*10*1) <= 0 else (float(reda[4])*10*1))  
        entry_temperature_data.append(float(reda[5])*2)
        time_data.append(float(reda[6])*1)
        event_data.append(str(reda[7]))
    step_data.append(float(roast_data_load[-3]))
    step_data.append(float(roast_data_load[-2]))
    step_data.append(float(roast_data_load[-1]))

    for j in range(len(time_data)-1):
        canvas.create_line(40+time_data[j],600-bt_temperature_data[j],40+time_data[j+1],600-bt_temperature_data[j+1],width=2,fill='red')
        canvas.create_line(40+time_data[j],600-ror_bt[j],40+time_data[j+1],600-ror_bt[j+1],width=2,fill='blue')
        canvas.create_line(40+time_data[j],600-et_temperature_data[j],40+time_data[j+1],600-et_temperature_data[j+1],width=2,fill='blue')
        canvas.create_line(40+time_data[j],600-ror_et[j],40+time_data[j+1],600-ror_et[j+1],width=2,fill='green')
        canvas.create_line(40+time_data[j],600-entry_temperature_data[j],40+time_data[j+1],600-entry_temperature_data[j+1],width=2,fill='brown')
    try:
        p_s = 40
        p_d = p_s + (step_data[0])*1
        p_m = p_s + (step_data[1])*1
        p_f = p_s + (step_data[2])*1
        pp_d = round(((step_data[0]) / (step_data[2]) * 100),2)
        pp_m = round((((step_data[1]) - (step_data[0])) / (step_data[2]) * 100),2)
        pp_f = round((((step_data[2]) - (step_data[1])) / (step_data[2]) * 100),2)

        canvas.create_line(p_s, 20, p_d, 20, width=8, fill='green')  #脫水時間
        canvas.create_line(p_d, 20, p_m, 20, width=8, fill='#FF8800')    #梅納時間
        canvas.create_line(p_m, 20, p_f, 20, width=8, fill='brown')  #發展時間

        canvas.create_text(p_s + 40,40,text=str(pp_d)+' %',fill='green')
        canvas.create_text(p_d + 40,40,text=str(pp_m)+' %',fill='#FF8800')
        canvas.create_text(p_m + 40,40,text=str(pp_f)+' %',fill='brown')
    except:
        messagebox.showinfo('歷史檔案', '不知道怎麼了，反正發生錯誤惹')

    #********** 繪圖區段 END ***********
    #********** 相關資訊區段 ***********

    #re_filename_L = Label(root_redraw_profile,text="檔案名稱",font="Keiu 10")#記錄檔名
    #re_filename_L.grid(row=12,column=0,padx=5,pady=5)
    #re_filename_E = Entry(root_redraw_profile,width=10,font="Keiu 10")
    #re_filename_E.grid(row=12,column=1,padx=5,pady=5)
    #re_filename_E.insert(0,'Test')#內定檔名

    re_roastdate_L = Label(root_redraw_profile,text="烘焙日期",font="Keiu 10")#烘焙日期
    re_roastdate_L.grid(row=12,column=2,padx=5,pady=5)
    re_roastdate_E = Entry(root_redraw_profile,width=20,font="Keiu 10")
    re_roastdate_E.grid(row=12,column=3,padx=5,pady=5)
    re_roastdate_E.insert(0,roast_data_load[0])#內定日期

    re_prodname_L = Label(root_redraw_profile,text="產品名稱",font="Keiu 10")#記錄產品名稱
    re_prodname_L.grid(row=12,column=4,padx=5,pady=5)
    re_prodname_E = Entry(root_redraw_profile,width=20,font="Keiu 10")
    re_prodname_E.grid(row=12,column=5,padx=5,pady=5)
    re_prodname_E.insert(0,roast_data_load[1])#內定產品名稱

    re_greenbean_inf_L = Label(root_redraw_profile,text="生豆資訊 : ",font="Keiu 10")#生豆資訊
    re_greenbean_inf_L.grid(row=13,column=0,padx=5,pady=5)

    re_gb_area_L = Label(root_redraw_profile,text="生豆處理法",font="Keiu 10")#生豆處理法
    re_gb_area_L.grid(row=13,column=1,padx=5,pady=5)
    re_gb_area_E = Entry(root_redraw_profile,width=20,font="Keiu 10")
    re_gb_area_E.grid(row=13,column=2,padx=5,pady=5)
    re_gb_area_E.insert(0,roast_data_load[2])#內定處理法 
       
    re_gb_name_L = Label(root_redraw_profile,text="生豆名稱",font="Keiu 10")#生豆名稱
    re_gb_name_L.grid(row=13,column=3,padx=5,pady=5)
    re_gb_name_E = Entry(root_redraw_profile,width=20,font="Keiu 10")
    re_gb_name_E.grid(row=13,column=4,padx=5,pady=5)
    re_gb_name_E.insert(0,roast_data_load[3])#內定生豆名稱

    re_gb_moisture_L = Label(root_redraw_profile,text="水分含量",font="Keiu 10")#水分含量
    re_gb_moisture_L.grid(row=13,column=5,padx=5,pady=5)
    re_gb_moisture_E = Entry(root_redraw_profile,width=20,font="Keiu 10")
    re_gb_moisture_E.grid(row=13,column=6,padx=5,pady=5)
    re_gb_moisture_E.insert(0,roast_data_load[4])#內定水分含量

    re_gb_density_L = Label(root_redraw_profile,text="生豆密度",font="Keiu 10")#生豆密度
    re_gb_density_L.grid(row=13,column=7,padx=5,pady=5)
    re_gb_density_E = Entry(root_redraw_profile,width=20,font="Keiu 10")
    re_gb_density_E.grid(row=13,column=8,padx=5,pady=5)
    re_gb_density_E.insert(0,roast_data_load[5])#內定生豆密度

    re_gb_weight_L = Label(root_redraw_profile,text="批次重量",font="Keiu 10")#批次重量
    re_gb_weight_L.grid(row=13,column=9,padx=5,pady=5)
    re_gb_weight_E = Entry(root_redraw_profile,width=20,font="Keiu 10")
    re_gb_weight_E.grid(row=13,column=10,padx=5,pady=5)
    re_gb_weight_E.insert(0,roast_data_load[6])#內定批次重量

    re_equipment_L = Label(root_redraw_profile,text="烘豆機資訊 : ",font="Keiu 10")#烘豆機資訊
    re_equipment_L.grid(row=14,column=0,padx=5,pady=5)

    re_machine_name_L = Label(root_redraw_profile,text="設備名稱",font="Keiu 10")#設備名稱
    re_machine_name_L.grid(row=14,column=1,padx=5,pady=5)
    re_machine_name_E = Entry(root_redraw_profile,width=20,font="Keiu 10")
    re_machine_name_E.grid(row=14,column=2,padx=5,pady=5)
    re_machine_name_E.insert(0,roast_data_load[7])#內定設備名稱 
       
    re_content_L = Label(root_redraw_profile,text="設備容量",font="Keiu 10")#設備容量
    re_content_L.grid(row=14,column=3,padx=5,pady=5)
    re_content_E = Entry(root_redraw_profile,width=20,font="Keiu 10")
    re_content_E.grid(row=14,column=4,padx=5,pady=5)
    re_content_E.insert(0,roast_data_load[8])#內定設備容量

    re_energy_L = Label(root_redraw_profile,text="使用能源",font="Keiu 10")#使用能源
    re_energy_L.grid(row=14,column=5,padx=5,pady=5)
    re_energy_E = Entry(root_redraw_profile,width=20,font="Keiu 10")
    re_energy_E.grid(row=14,column=6,padx=5,pady=5)
    re_energy_E.insert(0,roast_data_load[9])#內定使用能源

    re_operator_L = Label(root_redraw_profile,text="操作人員",font="Keiu 10")#操作人員
    re_operator_L.grid(row=14,column=7,padx=5,pady=5)
    re_operator_E = Entry(root_redraw_profile,width=20,font="Keiu 10")
    re_operator_E.grid(row=14,column=8,padx=5,pady=5)
    re_operator_E.insert(0,roast_data_load[10])#內定操作人員

    re_weather_L = Label(root_redraw_profile,text="天氣溫度",font="Keiu 10")#天氣溫度
    re_weather_L.grid(row=14,column=9,padx=5,pady=5)
    re_weather_E = Entry(root_redraw_profile,width=20,font="Keiu 10")
    re_weather_E.grid(row=14,column=10,padx=5,pady=5)
    re_weather_E.insert(0,roast_data_load[11])#內定天氣溫度

    re_cb_weight_L = Label(root_redraw_profile,text="熟豆重量",font="Keiu 10")#熟豆重量
    re_cb_weight_L.grid(row=15,column=1,padx=5,pady=5)
    re_cb_weight_E = Entry(root_redraw_profile,width=20,font="Keiu 10")
    re_cb_weight_E.grid(row=15,column=2,padx=5,pady=5)
    re_cb_weight_E.insert(0,roast_data_load[12])#內定熟豆重量

    re_loss_weight_L = Label(root_redraw_profile,text="失重率",font="Keiu 10")#失重率
    re_loss_weight_L.grid(row=15,column=3,padx=5,pady=5)
    re_loss_weight_E = Entry(root_redraw_profile,width=20,font="Keiu 10")
    re_loss_weight_E.grid(row=15,column=4,padx=5,pady=5)
    re_loss_weight_E.insert(0,roast_data_load[13])#內定失重率

    re_cb_barg_L = Label(root_redraw_profile,text="豆表色度",font="Keiu 10")#豆表色度
    re_cb_barg_L.grid(row=15,column=5,padx=5,pady=5)
    re_cb_barg_E = Entry(root_redraw_profile,width=20,font="Keiu 10")
    re_cb_barg_E.grid(row=15,column=6,padx=5,pady=5)
    re_cb_barg_E.insert(0,roast_data_load[14])#內定豆表色度

    re_cb_parg_L = Label(root_redraw_profile,text="豆粉色度",font="Keiu 10")#豆粉色度
    re_cb_parg_L.grid(row=15,column=7,padx=5,pady=5)
    re_cb_parg_E = Entry(root_redraw_profile,width=20,font="Keiu 10")
    re_cb_parg_E.grid(row=15,column=8,padx=5,pady=5)
    re_cb_parg_E.insert(0,roast_data_load[15])#內定豆粉色度

    re_cb_cupping_L = Label(root_redraw_profile,text="杯測風味",font="Keiu 10")#風味
    re_cb_cupping_L.grid(row=15,column=9,padx=5,pady=5)
    re_cb_cupping_E = Entry(root_redraw_profile,width=20,font="Keiu 10")
    re_cb_cupping_E.grid(row=15,column=10,padx=5,pady=5)
    re_cb_cupping_E.insert(0,roast_data_load[16])#內定風味

    #********** 相關資訊區段 End ***********

    re_su = Button(root_redraw_profile,text="確定",command=root_redraw_profile.destroy)
    re_su.grid(row=12,column=8,padx=5,pady=5)#,style='W.TButton'

    re_su = Button(root_redraw_profile,text="轉成 Artisan 格式",command=lambda:artisan_format(roast_data_load,r_f))
    re_su.grid(row=12,column=10,padx=5,pady=5)#,style='W.TButton'

    root_redraw_profile.mainloop()
    return

def load_file_data(filetype):#載入檔案
    inputfile = []
    root_ldf = Tk()
    
    root_ldf.geometry("800x100")
    var=StringVar()
    select_roast_file_cb = Combobox(root_ldf,width=60,textvariable=var,font="Keiu 14")#_bytesize
    select_roast_file_cb.grid(row=0,column=1,columnspan=4,pady=5,padx=5)

    if filetype == 1 or filetype ==2 :#載入歷史檔案
        #1:選擇已存資料  2:產生紀錄表
        root_ldf.title("歷史檔案選擇")
        for dirpath , dirnames, filenames in os.walk(os.getcwd()):
            for f in filenames:
                if f.split('.')[-1] == 'rxt' :#f
                   inputfile.append(os.path.join(dirpath,f))#os.path.join(dirpath,f)

        select_roast_file_cb["value"] = inputfile
        r_f = select_roast_file_cb

        Button(root_ldf,text="選擇檔案", style='W.TButton',command=lambda:argument_setup(4,r_f,filetype)).grid(row=0,column=8,padx=5,pady=5)
    elif filetype == 3 :#選擇已存通訊參數檔案
        root_ldf.title("檔案通訊參數選擇")
        for dirpath , dirnames, filenames in os.walk(os.getcwd()):
            for f in filenames:
                if f.split('.')[-1] == 'arg' :
                    inputfile.append(f)#os.path.join(dirpath,f)
  
        select_roast_file_cb["value"] = inputfile
        r_f = select_roast_file_cb
        Button(root_ldf,text="選擇檔案", style='W.TButton',command=lambda:argument_setup(3,r_f)).grid(row=0,column=8,padx=5,pady=5)
    elif filetype == 6 :#選擇Artisan 的.alog 檔案
        #將檔名傳入至argument_setup(6,r_f_as)
        #再將格式轉為rxt
        root_ldf.title("Artisan檔案選擇")
        for dirpath , dirnames, filenames in os.walk(os.getcwd()):
            for f in filenames:
                if f.split('.')[-1] == 'alog' :
                    inputfile.append(f)#os.path.join(dirpath,f)
  
        select_roast_file_cb["value"] = inputfile
        r_f = select_roast_file_cb
        Button(root_ldf,text="選擇檔案", style='W.TButton',command=lambda:argument_setup(6,r_f)).grid(row=0,column=8,padx=5,pady=5)

    Button(root_ldf,text="確定", style='W.TButton',command=root_ldf.destroy).grid(row=1,column=8,padx=5,pady=5)

    root_ldf.mainloop()
    return

def mouseMotion(event): #顯示滑鼠座標位置的時間溫度值
    x = event.x
    y = event.y
    x= (x - 66)*2
    x_time_min = x // 60
    x_time_sec = x % 60
    x_time = str(x_time_min) +':'+str(x_time_sec)

    y = round(((420 - y) / 1.6),2)
    y_temp = str(y) +' C'
    show_info = x_time + '   ' + y_temp
    mou_x_l.config(text=show_info)

    return
  
def frame3_modu():  #設備參數參考資料
    #參數標頭
    paramater_c1_l = Label(frame3, width = 12, background="white",foreground="blue",font="Helvetica 14 bold")
    paramater_c1_l.grid(row=23,column=1,pady=5,padx=5)
    paramater_c1_l.config(text='PV Address')
    paramater_c2_l = Label(frame3, width = 12, background="white",foreground="blue",font="Helvetica 14 bold")
    paramater_c2_l.grid(row=23,column=2,pady=5,padx=5)
    paramater_c2_l.config(text='SV Address')
    paramater_c3_l = Label(frame3, width = 12, background="white",foreground="blue",font="Helvetica 14 bold")
    paramater_c3_l.grid(row=23,column=3,pady=5,padx=5)
    paramater_c3_l.config(text='Run Stop Adr')
    #設備標頭
    delta_dtb_l = Label(frame3, width = 12, background="white",foreground="blue",font="Helvetica 14 bold")
    delta_dtb_l.grid(row=25,column=0,pady=5,padx=5)
    delta_dtb_l.config(text='Delta DTB')

    delta_dtk_l = Label(frame3, width = 12, background="white",foreground="blue",font="Helvetica 14 bold")
    delta_dtk_l.grid(row=27,column=0,pady=5,padx=5)
    delta_dtk_l.config(text='Delta DTK')

    omron_e5cc_l = Label(frame3, width = 12, background="white",foreground="blue",font="Helvetica 14 bold")
    omron_e5cc_l.grid(row=29,column=0,pady=5,padx=5)
    omron_e5cc_l.config(text='Omron e5cc')

    AAAAA_l = Label(frame3, width = 12, background="white",foreground="blue",font="Helvetica 14 bold")
    AAAAA_l.grid(row=31,column=0,pady=5,padx=5)
    AAAAA_l.config(text='AAAAA')

    BBBBB_l = Label(frame3, width = 12, background="white",foreground="blue",font="Helvetica 14 bold")
    BBBBB_l.grid(row=33,column=0,pady=5,padx=5)
    BBBBB_l.config(text='BBBBB')

    #參數內容
    delta_dtb_PV_l = Label(frame3, width = 12, background="yellow",foreground="blue",font="Helvetica 14 bold")
    delta_dtb_PV_l.grid(row=25,column=1,pady=5,padx=5)
    delta_dtb_PV_l.config(text='18176')
    delta_dtb_SV_l = Label(frame3, width = 12, background="yellow",foreground="blue",font="Helvetica 14 bold")
    delta_dtb_SV_l.grid(row=25,column=2,pady=5,padx=5)
    delta_dtb_SV_l.config(text='18177')
    delta_dtb_rs_l = Label(frame3, width = 12, background="yellow",foreground="blue",font="Helvetica 14 bold")
    delta_dtb_rs_l.grid(row=25,column=3,pady=5,padx=5)
    delta_dtb_rs_l.config(text='18201')

    delta_dtk_PV_l = Label(frame3, width = 12, background="yellow",foreground="blue",font="Helvetica 14 bold")
    delta_dtk_PV_l.grid(row=27,column=1,pady=5,padx=5)
    delta_dtk_PV_l.config(text='4096')
    delta_dtk_SV_l = Label(frame3, width = 12, background="yellow",foreground="blue",font="Helvetica 14 bold")
    delta_dtk_SV_l.grid(row=27,column=2,pady=5,padx=5)
    delta_dtk_SV_l.config(text='4097')
    delta_dtk_rs_l = Label(frame3, width = 12, background="yellow",foreground="blue",font="Helvetica 14 bold")
    delta_dtk_rs_l.grid(row=27,column=3,pady=5,padx=5)
    delta_dtk_rs_l.config(text='4120')

    omron_e5cc_PV_l = Label(frame3, width = 12, background="lightgreen",foreground="red",font="Helvetica 14 bold")
    omron_e5cc_PV_l.grid(row=29,column=1,pady=5,padx=5)
    omron_e5cc_PV_l.config(text='8192')
    omron_e5cc_SV_l = Label(frame3, width = 12, background="lightgreen",foreground="red",font="Helvetica 14 bold")
    omron_e5cc_SV_l.grid(row=29,column=2,pady=5,padx=5)
    omron_e5cc_SV_l.config(text='8451')
    omron_e5cc_rs_l = Label(frame3, width = 12, background="lightgreen",foreground="red",font="Helvetica 14 bold")
    omron_e5cc_rs_l.grid(row=29,column=3,pady=5,padx=5)
    omron_e5cc_rs_l.config(text='nnnn')

    AAAAA_PV_l = Label(frame3, width = 12, background="lightgreen",foreground="red",font="Helvetica 14 bold")
    AAAAA_PV_l.grid(row=31,column=1,pady=5,padx=5)
    AAAAA_PV_l.config(text='0000')
    AAAAA_SV_l = Label(frame3, width = 12, background="lightgreen",foreground="red",font="Helvetica 14 bold")
    AAAAA_SV_l.grid(row=31,column=2,pady=5,padx=5)
    AAAAA_SV_l.config(text='0000')
    AAAAA_rs_l = Label(frame3, width = 12, background="lightgreen",foreground="red",font="Helvetica 14 bold")
    AAAAA_rs_l.grid(row=31,column=3,pady=5,padx=5)
    AAAAA_rs_l.config(text='nnnn')

    BBBBB_PV_l = Label(frame3, width = 12, background="lightgreen",foreground="red",font="Helvetica 14 bold")
    BBBBB_PV_l.grid(row=33,column=1,pady=5,padx=5)
    BBBBB_PV_l.config(text='0000')
    BBBBB_SV_l = Label(frame3, width = 12, background="lightgreen",foreground="red",font="Helvetica 14 bold")
    BBBBB_SV_l.grid(row=33,column=2,pady=5,padx=5)
    BBBBB_SV_l.config(text='0000')
    BBBBB_rs_l = Label(frame3, width = 12, background="lightgreen",foreground="red",font="Helvetica 14 bold")
    BBBBB_rs_l.grid(row=33,column=3,pady=5,padx=5)
    BBBBB_rs_l.config(text='nnnn')
    return

def frame4_modu():  #烘豆手法選擇
    var_rostype = IntVar()
    var_rostype.set(2)
    rostype_nd_1 = Radiobutton(frame4,text="北歐烘焙法",variable=var_rostype,value = 1,command=lambda:draw_panal_ss(1))
    rostype_nd_1.grid(row=0,column=0,columnspan=14,padx=5,pady=5)

    rostype_sr_2 = Radiobutton(frame4,text="Scott Rao_漸降式烘焙法",variable=var_rostype,value = 2,command=lambda:draw_panal_ss(2))
    rostype_sr_2.grid(row=0,column=15,columnspan=14,padx=5,pady=5)

    rostype_sn_3 = Radiobutton(frame4,text="小野善造_完全烘焙法",variable=var_rostype,value = 3,command=lambda:draw_panal_ss(3))
    rostype_sn_3.grid(row=15,column=0,columnspan=14,padx=5,pady=5)

    rostype_ss_4 = Radiobutton(frame4,text="全息烘焙法",variable=var_rostype,value = 4,command=lambda:draw_panal_ss(4))
    rostype_ss_4.grid(row=15,column=15,columnspan=14,padx=5,pady=5)
    #******************************************************************************************
    #北歐烘焙法
    rostype_nd_text= Text(frame4,height=15,width=65 ,foreground='black',font="Keiu 16")
    rostype_nd_text.grid(row=2,column=0,columnspan=14,padx=5,pady=5)
    rostype_nd_text.insert(END,"入豆溫接近一爆溫(風門全開)，回溫點約在40 ~ 60秒\n")
    rostype_nd_text.insert(END,"一爆起26 ~ 120秒之間下豆，義式豆一爆後120 ~ 180秒下豆\n")
    rostype_nd_text.insert(END,"總烘焙時間約在9分半至10分，常態上於一爆密集下豆\n")
    rostype_nd_text.insert(END,"磨粉後內外差20以上(常見 外75 / 內100)\n")
    rostype_nd_text.insert(END,"適合水果調性豐富的生豆，如肯亞SL28/SL34，衣索比亞原生種，巴拿馬藝妓\n")

    #分隔線
    line_4_1 = ttk.Separator(frame4, orient=VERTICAL, style="Line.TSeparator").grid(row=0, rowspan=20,column=14, sticky='ns')

    #Scott Rao_漸降式烘焙法
    rostype_sr_text= Text(frame4,height=15,width=65, foreground='blue',font="Keiu 16")
    rostype_sr_text.grid(row=2,column=15,columnspan=14,padx=5,pady=5)
    rostype_sr_text.insert(END,"1.烘焙前期給予較大的火力，能夠幫助整個烘焙環境迅速回到合理範圍。\n")
    rostype_sr_text.insert(END,"2.如果採用前段悶蒸法(Soak)，建議第一分鐘的火力維持20%，而通常\n")
    rostype_sr_text.insert(END,"使用悶蒸法的時機在於用12公斤以下的小型烘豆機，或是鍋次之間烘焙\n")
    rostype_sr_text.insert(END,"機的溫度無法平順下降時使用。\n")
    rostype_sr_text.insert(END,"3.火力建議能在一爆前40 - 45秒之間進行調整。在調整瓦斯之後，如果\n")
    rostype_sr_text.insert(END,"ROR在一爆前後出現變平、上升又失速，代表火力太大；\n")
    rostype_sr_text.insert(END,"如果一爆前ROR變成平坦然後失速，代表火力過低。\n")
    rostype_sr_text.insert(END,"4.一爆開始後直到發展率(DTR)占整體發展時間12%之前，都是失速危險\n")
    rostype_sr_text.insert(END,"期。而當一爆後發展率(DTR)超過16%，則是升溫暴衝的危險期。\n")
    rostype_sr_text.insert(END,"因此，一爆後最佳調降火力的時機，是在發展率(DTR) 12% ~ 16%之間\n")
    rostype_sr_text.insert(END,"5.入風溫室一個有用的指標，用來預測及調整烘焙曲線，\n")
    rostype_sr_text.insert(END,"而與火力大小有較大的關聯性，可藉由火力調整來影響入風溫。\n")
    rostype_sr_text.insert(END,"\n")

    #橫分隔線
    line_4_2 = ttk.Separator(frame4, orient=HORIZONTAL, style="Line.TSeparator").grid(row=14,column=0, columnspan=30, sticky='ew')#

    #小野善造_完全烘焙法
    rostype_sn_text= Text(frame4,height=15,width=65, foreground='red',font="Keiu 16")
    rostype_sn_text.grid(row=17,column=0,columnspan=14,padx=5,pady=5)
    rostype_sn_text.insert(END,"1.又稱'長時間烘焙'，其中分為三段:\n")
    rostype_sn_text.insert(END,"*'預備烘焙':豆溫0 ~ 130度約第7分鐘，到達160度約第12分鐘\n")
    rostype_sn_text.insert(END,"*'正式烘焙':豆溫160至185度之間，此時每分鐘ROR約4~5度(以1公斤烘焙為例)；\n")
    rostype_sn_text.insert(END,"一爆點(185度)約17 ~ 19分鐘；二爆點208度約在第23分鐘，在豆溫\n")
    rostype_sn_text.insert(END,"218度約第26分鐘時結束二爆。\n")
    rostype_sn_text.insert(END,"*'最終烘焙':使溫度停止變化，並維持固定的溫度。用來修正豆子的發展狀態。\n")
    rostype_sn_text.insert(END,"2.一爆/二爆皆是指'開始連續爆裂'之後算起。\n")
    rostype_sn_text.insert(END,"3.重視風門的操作，認為是烘焙過程中的首要條件，也是決定烘焙成果優劣的關鍵。\n")
    rostype_sn_text.insert(END,"4.醇厚度、甜味、苦味的均衡。利於保存，風味衰敗速度慢。\n")
    rostype_sn_text.insert(END,"\n")

    #全息烘焙法
    rostype_ss_text= Text(frame4,height=15,width=65, foreground='green',font="Keiu 16")
    rostype_ss_text.grid(row=17,column=15,columnspan=14,padx=5,pady=5)
    rostype_ss_text.insert(END,"第一階段:入豆到 T0 - T1 (玻璃態)\n")
    rostype_ss_text.insert(END,"    T0:豆表溫約在110°C\n")
    rostype_ss_text.insert(END,"第二階段:T1 至大理石紋 (橡膠態)\n")
    rostype_ss_text.insert(END,"    T1:約135 ~ 145°C\n")
    rostype_ss_text.insert(END,"第三階段:大理石紋至一爆(豆表進入到玻璃態)\n")
    rostype_ss_text.insert(END,"    大理石紋:約175 ~ 190°C\n")
    rostype_ss_text.insert(END,"第四階段:一爆至出鍋。\n")
    rostype_ss_text.insert(END,"\n")

    return

def step_change():  #重設階段
    canvas.create_rectangle(0,0, 900, 420,fill='#ffffff', stipple="gray50")
    canvas_ss.create_rectangle(0,0, 900, 420,fill='#ffffff', stipple="gray50")
    draw_panal()    #重繪繪圖座標系統
    draw_panal_ss(0)    #重繪繪圖座標系統
    return

def artisan_format(roast_data_load,r_f):#將rxt檔案轉成csv格式
    #.rxt原始存檔順序-->序號:豆溫:豆溫ror:環境溫:環境溫ror:入風溫:時間:事件
    w_artisan_Time1 = []
    w_artisan_Time2 = []
    w_artisan_ET = []
    w_artisan_BT = []
    w_artisan_Event = []
    w_Event = ""
    artisan_CHARGE = '' ; artisan_TP = '' ; artisan_DRYe = ''
    artisan_FCs = '' ; artisan_FCe = '' ; artisan_SCs = '' ; artisan_SCe = '' 
    artisan_DROP = '' 
    #將日期格式改為Artisan的格式:dd.mm.yyyy
    artisan_date = roast_data_load[0][-3:-1]+'.'+roast_data_load[0][-6:-4]+'.'+roast_data_load[0][0:4]
    #print (artisan_date)
    #判斷事件的時間點並格式化時間且轉成Artisan支援的CSV檔案格式
    for k in range(17,len(roast_data_load)-3):
        reda = roast_data_load[k].split(':')
        if reda[7] == "CHARGE\n" :
            min_s = float(reda[6]) // 60
            sec_s = float(reda[6]) % 60
            data_s = str("{:02.0f}".format(min_s)) +':'+str("{:02.0f}".format(round(sec_s,1)))
            artisan_CHARGE = data_s
            w_Event = "CHARGE\n"
        elif reda[7] == "TP\n" :
            min_s = float(reda[6]) // 60
            sec_s = float(reda[6]) % 60
            data_s = str("{:02.0f}".format(min_s)) +':'+str("{:02.0f}".format(round(sec_s,1)))
            artisan_TP = data_s
            w_Event = "TP\n"
        elif reda[7] == "DRYe\n" :
            min_s = float(reda[6]) // 60
            sec_s = float(reda[6]) % 60
            data_s = str("{:02.0f}".format(min_s)) +':'+str("{:02.0f}".format(round(sec_s,1)))
            artisan_DRYe = data_s
            w_Event = "DRYe\n"
        elif reda[7] == "FCs\n" :
            min_s = float(reda[6]) // 60
            sec_s = float(reda[6]) % 60
            data_s = str("{:02.0f}".format(min_s)) +':'+str("{:02.0f}".format(round(sec_s,1)))
            artisan_FCs = data_s
            w_Event = "FCs\n"
        elif reda[7] == "FCe\n" :
            min_s = float(reda[6]) // 60
            sec_s = float(reda[6]) % 60
            data_s = str("{:02.0f}".format(min_s)) +':'+str("{:02.0f}".format(round(sec_s,1)))
            artisan_FCe = data_s
            w_Event = "FCe\n"
        elif reda[7] == "SCs\n" :
            min_s = float(reda[6]) // 60
            sec_s = float(reda[6]) % 60
            data_s = str("{:02.0f}".format(min_s)) +':'+str("{:02.0f}".format(round(sec_s,1)))
            artisan_SCs = data_s
            w_Event = "SCs\n"
        elif reda[7] == "SCe\n" :
            min_s = float(reda[6]) // 60
            sec_s = float(reda[6]) % 60
            data_s = str("{:02.0f}".format(min_s)) +':'+str("{:02.0f}".format(round(sec_s,1)))
            artisan_SCe = data_s
            w_Event = "SCe\n"
        elif reda[7] == "DROP\n" :
            min_s = float(reda[6]) // 60
            sec_s = float(reda[6]) % 60
            data_s = str("{:02.0f}".format(min_s)) +':'+str("{:02.0f}".format(round(sec_s,1)))
            artisan_DROP = data_s            
            w_Event = "DROP\n"
        elif reda[7] == "金黃點\n" :
            min_s = float(reda[6]) // 60
            sec_s = float(reda[6]) % 60
            data_s = str("{:02.0f}".format(min_s)) +':'+str("{:02.0f}".format(round(sec_s,1)))
            artisan_GDp = data_s            
            w_Event = "\n"
        else:
            #w_Event = reda[7]
            w_Event = '\n'

        w_artisan_Time1.append(reda[0])
        w_artisan_Time2.append(reda[6])
        w_artisan_ET.append(reda[3])
        w_artisan_BT.append(reda[1])
        w_artisan_Event.append(w_Event)
        w_Event = ""

    artisan_COOL = ''
    min_s = float(reda[6]) // 60
    sec_s = float(reda[6]) % 60
    data_s = str("{:02.0f}".format(min_s)) +':'+str("{:02.0f}".format(round(sec_s,1)))
    artisan_Time = data_s

    with open( r_f +'.csv','w') as f:

        f.write("Date:"+artisan_date+"	")  #烘焙日期
        f.write("Unit:C"+"	")  #Unit
        f.write("CHARGE:"+artisan_CHARGE+"	")  #烘焙日期
        f.write("TP:"+artisan_TP+"	")  #回溫點
        f.write("DRYe:"+artisan_DRYe+"	")  #
        f.write("FCs:"+artisan_FCs+"	")  #一爆開始
        f.write("FCe:"+artisan_FCe+"	")  #一爆結束
        f.write("SCs:"+artisan_SCs+"	")  #二爆開始
        f.write("SCe:"+artisan_SCe+"	")  #二爆結束
        f.write("DROP:"+artisan_DROP+"	")  #下豆
        f.write("COOL:"+artisan_COOL+"	")  #冷卻
        f.write("Time:"+artisan_Time)  #烘焙時間
        f.write('\n')  #
        f.write("Time1"+"	"+"Time2"+"	"+"ET"+"	"+"BT"+"	"+"Event")
        f.write('\n')  #

    #--------將資料轉成Artisan支援的CSV檔案格式並寫入檔案-------
        for h in range(len(w_artisan_Time1)):
            time1_min = float(w_artisan_Time1[h]) // 60
            time1_sec = float(w_artisan_Time1[h]) % 60
            time1_data = str("{:02.0f}".format(time1_min)) +':'+str("{:02.0f}".format(round(time1_sec,1)))
            #time1_data = str(time1_min) +':'+str(round(time1_sec,1))
            time2_min = float(w_artisan_Time2[h]) // 60
            time2_sec = float(w_artisan_Time2[h]) % 60
            #time2_data = str(time2_min) +':'+str(round(time2_sec,1))
            time2_data = str("{:02.0f}".format(time2_min)) +':'+str("{:02.0f}".format(round(time2_sec,1)))

            f.write(time1_data+"	"+time2_data+"	"+w_artisan_ET[h]+"	"+w_artisan_BT[h]+"	"+w_artisan_Event[h])#+"\n"
    messagebox.showinfo('OK', '轉換完成檔名為'+r_f +'.csv')
    return

def record_table(roast_data_load,r_f): #轉成烘焙紀錄表

    time_data = []
    bt_temperature_data =[]
    ror_bt = []
    et_temperature_data =[]
    ror_et = []
    entry_temperature_data =[]
    event_data = []
    step_data = []
    for i in range(17,len(roast_data_load)-3):
        reda = roast_data_load[i].split(':')
        #print(reda)
        bt_temperature_data.append(float(reda[1])*1)
        ror_bt.append(0 if (float(reda[2])*10*1) <= 0 else (float(reda[2])*10*1))
        et_temperature_data.append(float(reda[3])*1)
        ror_et.append(0 if (float(reda[4])*10*1) <= 0 else (float(reda[4])*10*1))  
        entry_temperature_data.append(float(reda[5])*1)
        time_data.append(float(reda[6])*1)
        event_data.append(str(reda[7]))
    step_data.append(float(roast_data_load[-3]))
    step_data.append(float(roast_data_load[-2]))
    step_data.append(float(roast_data_load[-1]))
    #print(time_data)
    wb = openpyxl.load_workbook('烘焙紀錄表.xlsx')
    sheet = wb.active
    sheet.cell(row=1,column=18).value = roast_data_load[10]   #'R1'操作人員
    sheet.cell(row=2,column=2).value = roast_data_load[0]   #'B2'烘焙日期
    sheet.cell(row=2,column=5).value = roast_data_load[11]   #'E2'天氣溫度
    sheet.cell(row=2,column=7).value = '室外溫度1'   #'G2'
    sheet.cell(row=2,column=9).value = '室內溫度1'   #'I2'
    sheet.cell(row=2,column=11).value = '濕度1'   #'K2'
    sheet.cell(row=2,column=12).value = '第    鍋'   #'L2'
    sheet.cell(row=2,column=15).value = roast_data_load[3]   #'O2'生豆名稱
    sheet.cell(row=3,column=2).value = roast_data_load[1]   #'B3'產區
    sheet.cell(row=3,column=5).value = roast_data_load[2]   #'E3'處理法
    sheet.cell(row=4,column=2).value = roast_data_load[6]   #'B4'生豆重量
    sheet.cell(row=4,column=4).value = roast_data_load[4]   #'D4'生豆含水率
    sheet.cell(row=5,column=2).value = roast_data_load[12]   #'B5'熟豆重量
    sheet.cell(row=5,column=4).value = '熟豆含水率1'   #'D5'
    sheet.cell(row=6,column=6).value = 'F6'   #'F6'
    sheet.cell(row=7,column=4).value = roast_data_load[14]   #'D7'烘焙度外
    sheet.cell(row=7,column=5).value = roast_data_load[15]   #'E7'烘焙度內
    sheet.cell(row=7,column=6).value = 'F7'   #'F7'
    sheet.cell(row=11,column=6).value = 'F11'   #'F11'
    sheet.cell(row=12,column=2).value = roast_data_load[16]   #'B12'杯測風味

    for k in range(len(time_data)):
        tp_index_1 = int(k) // 60
        tp_index_2 = int(k) % 60
        #print (tp_index_1,tp_index_2,time_data[k],bt_temperature_data[k])
        if (tp_index_2 == 59) :
            sheet.cell(row=15,column=tp_index_1+2).value = bt_temperature_data[k]

        if event_data[k] == 'CHARGE\n':
            sheet.cell(row=8,column=2).value = bt_temperature_data[k]   #'B8'進豆溫度
        elif event_data[k] == 'TP\n':
            sheet.cell(row=8,column=4).value = bt_temperature_data[k]   #'D8'回溫點
            time_min = time_data[k] // 60
            time_sec = time_data[k] % 60
            show_time = str(time_min) +':'+str(time_sec)+''
            sheet.cell(row=8,column=6).value = show_time   #'F8'回溫時間        
        elif event_data[k] == 'DRYe\n':
            time_min = time_data[k] // 60
            time_sec = time_data[k] % 60
            show_time = str(time_min) +':'+str(time_sec)+''
            sheet.cell(row=4,column=6).value = show_time   #'F4'轉黃點 150"C
        elif event_data[k] == 'GDp\n':
            time_min = time_data[k] // 60
            time_sec = time_data[k] % 60
            show_time = str(time_min) +':'+str(time_sec)+''
            sheet.cell(row=5,column=6).value = show_time   #'F5'金黃點 170"C
        elif event_data[k] == 'FCs\n':
            sheet.cell(row=9,column=2).value = bt_temperature_data[k]   #'B9'一爆溫度
            time_min = time_data[k] // 60
            time_sec = time_data[k] % 60
            show_time = str(time_min) +':'+str(time_sec)+''
            sheet.cell(row=9,column=4).value = show_time   #'D9'一爆時間
            sheet.cell(row=9,column=6).value = ''   #'F9'一爆送風
        elif event_data[k] == 'SCs\n':
            sheet.cell(row=10,column=2).value = bt_temperature_data[k]   #'B10'二爆溫度
            time_min = time_data[k] // 60
            time_sec = time_data[k] % 60
            show_time = str(time_min) +':'+str(time_sec)+''
            sheet.cell(row=10,column=4).value = show_time   #'D10'二爆時間
            sheet.cell(row=10,column=6).value = ''   #'F10'二爆送風
        elif event_data[k] == 'DROP\n':
            sheet.cell(row=11,column=2).value = bt_temperature_data[k]   #'B11'下豆溫度
            time_min = time_data[k] // 60
            time_sec = time_data[k] % 60
            show_time = str(time_min) +':'+str(time_sec)+''
            sheet.cell(row=11,column=4).value = show_time   #'D11'下豆時間

    sheet.cell(row=19,column=2).value = '備註1'   #'B19'
    #print(sheet.cell(row=2,column=1).value)
    wb.save(r_f + '.xlsx')#r_f
    messagebox.showinfo('烘焙紀錄表', '烘焙紀錄表轉換完成 ' + r_f +'.xlsx')
    return

if __name__ == '__main__' :#主程式及使用者介面設定
    matplotlib.use('TkAgg')
    global state_a #紀錄烘豆事件即離開溫度記錄的變數
    root_t = Tk()
    screenwidth = root_t.winfo_screenwidth()
    screenheight = root_t.winfo_screenheight()
    w_win = 1480
    h_win = 760
    x_offset = (screenwidth - w_win) / 2
    y_offset = ((screenheight - h_win) / 2)# - 30

    root_t.title("Shokunin Coffee Roaster")
    root_t.geometry("%dx%d+%d-%d" %(w_win,h_win,x_offset,y_offset))#1520
    root_t.configure(bg='lightblue')
    root_t.iconbitmap("coffee_logo.ico")
    
    st = Style() #設定按鈕外觀
    st.configure('W.TButton', width = 10, background='green', foreground='blue', font=('Keiu', 14 ))
    st.configure('W1.TButton', width = 10, background='#FF8000', foreground='red', font=('Keiu', 14 ))
    st.configure('W2.TButton', width = 10,background='grey', foreground='#aa0011', font=('Keiu', 14 ))
    st.configure('W3.TButton', width = 14, background='#FF8000', foreground='blue', font=('Keiu', 14 ))
    st.configure('Wf.TButton', width = 6,background='#7CFC00', foreground='blue', font=('Keiu', 14 ))
    st.configure('E1.TEntry',background='#7CFC00', foreground='red', font=('Keiu', 16 ))

    ttk.Style().configure(".", font=('Keiu', 12)) # notebook標籤字體  "."
    ttk.Style().configure("Line.TSeparator", background="#ff0011")##ff2266
    ttk.Style().configure('.', foreground='black')
    notebook = ttk.Notebook(root_t)
    #*************** 第一個視窗-繪圖 BT & ROR ***************
    frame1 = Frame (root_t, relief=RIDGE, borderwidth=2)#GROOVE RAISED RIDGE
    frame1.pack(side=BOTTOM, fill='both', expand=1)#, ipadx="1c", ipady="1c"
    
    #在Tk的GUI上放置一個畫布，並用.grid()來調整佈局

    canvas =Canvas(frame1, width=940, height=460, bg='white')
    canvas.grid(row=0, rowspan=14 ,column=2,columnspan=6, padx=0, pady=0)
    canvas_event =Canvas(frame1, width=940, height=130, bg='white')
    canvas_event.grid(row=14, rowspan=3 ,column=2,columnspan=6, padx=0, pady=0)
   
    #烘豆事件紀錄按鈕
    bt_charge = Button(frame1,text='入豆',style='W.TButton',state=DISABLED, command=lambda:roast_state('CHARGE'))
    bt_charge.grid(row=8,column=0)
    bt_tp = Button(frame1,text='回溫點',style='W.TButton',state=DISABLED, command=lambda:roast_state('TP'))
    bt_tp.grid(row=9,column=0)
    bt_drye = Button(frame1,text='脫水結束', style='W.TButton',state=DISABLED,command=lambda:roast_state('DRYe'))
    bt_drye.grid(row=10,column=0)
    bt_gdp = Button(frame1,text='金黃點', style='W1.TButton',state=DISABLED,command=lambda:roast_state('GDp'))
    bt_gdp.grid(row=11,column=0)
    bt_fcs = Button(frame1,text='一爆', style='W1.TButton',command=lambda:roast_state('FCs'))
    bt_fcs.grid(row=12,column=0)
    bt_fce = Button(frame1,text='一爆結束', style='W1.TButton',command=lambda:roast_state('FCe'))
    bt_fce.grid(row=13,column=0)
    bt_scs = Button(frame1,text='二爆', style='W2.TButton',command=lambda:roast_state('SCs'))
    bt_scs.grid(row=14,column=0)
    bt_sce = Button(frame1,text='二爆結束', style='W2.TButton',command=lambda:roast_state('SCe'))
    bt_sce.grid(row=15,column=0)
    bt_drop = Button(frame1,text='下豆', style='W2.TButton',command=lambda:roast_state('DROP'))
    bt_drop.grid(row=16,column=0)
    #********* 事件溫度時間顯示 *********
    entry_width = 12
    charge_E = Entry(frame1,style='E1.TEntry', width = entry_width)#
    charge_E.grid(row=8,column=1)
    charge_E.insert(0,'')
    rtp_E = Entry(frame1, width = entry_width)
    rtp_E.grid(row=9,column=1)
    rtp_E.insert(0,'')
    enddry_E = Entry(frame1, width = entry_width)
    enddry_E.grid(row=10,column=1)
    enddry_E.insert(0,'')
    gp_E = Entry(frame1, width = entry_width)
    gp_E.grid(row=11,column=1)
    gp_E.insert(0,'')
    fc_E = Entry(frame1, width = entry_width)
    fc_E.grid(row=12,column=1)
    fc_E.insert(0,'')
    fcend_E = Entry(frame1, width = entry_width)
    fcend_E.grid(row=13,column=1)
    fcend_E.insert(0,'')
    sc_E = Entry(frame1, width = entry_width)
    sc_E.grid(row=14,column=1)
    sc_E.insert(0,'')
    scend_E = Entry(frame1, width = entry_width)
    scend_E.grid(row=15,column=1)
    scend_E.insert(0,'')
    drop_E = Entry(frame1, width = entry_width)
    drop_E.grid(row=16,column=1)
    drop_E.insert(0,'')
    #********* 事件溫度時間顯示 end *********

    #烘焙計時器
    RoT_l = Label(frame1, text="Timer", width = 8, background="white", foreground="Green", font="Helvetica 10 bold").grid(row=0,column=8,pady=5,padx=5)
    RoT_2 = Label(frame1, background="white", foreground="blue", font="Helvetica 16 bold")
    RoT_2.grid(row=0,column=9,pady=5,padx=5)
 
    #*****----- 溫度 及 ROR -----*****
    #環境溫
    mark_et = Label(frame1,text='ET', width = 8, background="white",foreground="blue",font="Helvetica 10 bold").grid(row=2,column=8,pady=5,padx=5)
    et_temp_l = Label(frame1, width = 8, background="lightgreen",foreground="blue",font="Helvetica 16 bold")
    et_temp_l.grid(row=2,column=9,pady=5,padx=5)
    #豆溫
    mark_bt_l = Label(frame1,text='BT', width = 8, background="white",foreground="red",font="Helvetica 10 bold").grid(row=3,column=8,pady=5,padx=5)
    bt_temp_l = Label(frame1, width = 8, background="yellow",foreground="red",font="Helvetica 16 bold")
    bt_temp_l.grid(row=3,column=9,pady=5,padx=5)
    #入風溫
    mark_entry_l = Label(frame1,text='Entry', width = 8, background="lightblue",foreground="brown",font="Helvetica 10 bold").grid(row=4,column=8,pady=5,padx=5)
    entry_temp_l = Label(frame1, width = 8, background="lightblue",foreground="brown",font="Helvetica 16 bold")
    entry_temp_l.grid(row=4,column=9,pady=5,padx=5)
    #ET ROR
    #mark_et_ror = Label(frame1,text='△ET', width = 8, background="white",foreground="blue",font="Helvetica 10 bold").grid(row=5,column=8,pady=5,padx=5)
    #et_temp_ror_l = Label(frame1, width = 8, background="lightgreen",foreground="blue",font="Helvetica 16 bold")
    #et_temp_ror_l.grid(row=5,column=9,pady=5,padx=5)
    #BT ROR
    mark_bt_ror = Label(frame1,text='△BT', width = 8, background="white",foreground="green",font="Helvetica 10 bold").grid(row=6,column=8,pady=5,padx=5)
    bt_temp_ror_l = Label(frame1, width = 8, background="yellow",foreground="green",font="Helvetica 16 bold")
    bt_temp_ror_l.grid(row=6,column=9,pady=5,padx=5)

    #發展狀態顯示
    even_dev = Label(frame1,text='發展時間', width = 8, background="white",foreground="red",font="Helvetica 12 bold").grid(row=15,column=8,pady=5,padx=5)
    even_dev_l = Label(frame1, width = 8, background="yellow",foreground="red",font="Helvetica 12 bold")
    even_dev_l.grid(row=16,column=8,pady=5,padx=5)
    even_dev_p = Label(frame1,text='發展率', width = 8, background="white",foreground="brown",font="Helvetica 12 bold").grid(row=15,column=9,pady=5,padx=5)
    even_dev_p_l = Label(frame1, width = 8, background="green",foreground="yellow",font="Helvetica 12 bold")
    even_dev_p_l.grid(row=16,column=9,pady=5,padx=5)

    #壓差紀錄滑條
    fan_l = Label(frame1,width = 10 ,text="壓差計", background="white",foreground="blue",font="Helvetica 14 bold")
    fan_l.grid(row=18,column=2,pady=1,padx=1)
    slider = Scale(frame1, from_=0, to=60,length=800, orient='horizontal',command=sl_f_ch)#orient='vertical'
    slider.set(10)
    slider.grid(row=18,column=3,columnspan=2)
    slf_ch_b=Button(frame1,text='F10',style='Wf.TButton', command=lambda:roast_state('F10')).grid(row=18,column=5)

    #瓦斯紀錄滑條
    power_l = Label(frame1,width = 10,text="瓦斯KPa", background="white",foreground="red",font="Helvetica 14 bold")
    power_l.grid(row=19,column=2,pady=1,padx=1)
    slider_p = Scale(frame1, from_=0, to=100,length=800, orient='horizontal',command=sl_p_ch)
    slider_p.set(3)
    slider_p.grid(row=19, column=3,columnspan=2)
    slP_ch_b=Button(frame1,text='P3.0',style='Wf.TButton', command=lambda:roast_state('P3.0')).grid(row=19,column=5)

    #預計脫水結束時間
    even_predryend = Label(frame1,text='Pre DryEnd Time', width = 16, background="white",foreground="red",font="Helvetica 10 bold").grid(row=17,column=0,pady=5,padx=5)
    even_predryend_l = Label(frame1, width = 10, background="pink",foreground="red",font="Helvetica 14 bold")
    even_predryend_l.grid(row=17,column=1,pady=5,padx=5)

    #Logo插圖
    img = Image.open('coffee roast.jpg')
    img=img.resize((int(img.size[0]/2),int(img.size[1]/2)))
    tk_img = ImageTk.PhotoImage(img)
    canvas_img = Canvas(frame1,width=img.size[0],height=img.size[1])#width=img.size[0],height=img.size[1]
    canvas_img.create_image(0, 0, anchor=NW, image=tk_img)#tk_img'nw'
    canvas_img.grid(row=8,column=8,rowspan=7,columnspan=2,pady=5,padx=5)#,rowspan=15,columnspan=2
    
    #艾格狀數預估
    f1_agtron_L = Label(frame1,text="豆表 Agtron 預估值",font="Keiu 14")#當下艾格狀數預估
    f1_agtron_L.grid(row=17,column=8,columnspan=2,padx=5,pady=5)
    f1_agtron_L_S = Label(frame1,text="",foreground="red",font="Keiu 14")#FC 豆表 Agtron
    f1_agtron_L_S.grid(row=18,column=9,columnspan=2,padx=5,pady=5)

    #滑鼠座標顯示
    canvas.bind("<Motion>",mouseMotion)
    mou_x_l = Label(frame1, width = 14,foreground="blue",font="Helvetica 11 bold")
    mou_x_l.grid(row=0,column=0,columnspan=2,pady=5,padx=5)
    
    #開始、結束按鈕
    bt_openmc = Button(frame1,text="開機",state=DISABLED, style='W.TButton',command=lambda:temp_ror('0'))
    bt_openmc.grid(row=2,column=0,columnspan=2,padx=5,pady=5)    
    bt_records = Button(frame1,text="開使記錄",state=DISABLED, style='W.TButton',command=lambda:roast_state('go'))
    bt_records.grid(row=3,column=0,columnspan=2,padx=5,pady=5)
    bt_recorde = Button(frame1,text='結束記錄',state=DISABLED, style='W.TButton',command=lambda:temp_ror(1))
    bt_recorde.grid(row=4,column=0,columnspan=2,padx=5,pady=5)
    bt_cleantr = Button(frame1,text="清除表格",state=DISABLED, style='W.TButton',command=clean_tree)
    bt_cleantr.grid(row=5,column=0,columnspan=2,padx=5,pady=5)
    bt_end = Button(frame1,text="結束程式", style='W.TButton',command=root_t.destroy)
    bt_end.grid(row=6,column=0,columnspan=2,padx=5,pady=5)
  
    #*************** 第二個視窗-通訊參數 ***************
    frame2 = Frame (root_t, relief='solid', cursor='cross', borderwidth=2)
    frame2.pack (side='top', fill='both', ipadx="1c", ipady="1c", expand=1)

    #-----控制器通用參數------------------------------------------------------------------------
    port_L = Label(frame2,text="通信端口",font="Keiu 16")
    port_L.grid(row=0,column=0,pady=5)
    port_E = Entry(frame2,width=8,font="Keiu 16")
    port_E.grid(row=0,column=1,pady=5)
    port_E.insert(0,'COM5')

    mode_L = Label(frame2,text="通訊類型",font="Keiu 16")
    mode_L.grid(row=1,column=0,pady=5)
    mode_E = Entry(frame2,width=8,font="Keiu 16")
    mode_E.grid(row=1,column=1,pady=5)
    mode_E.insert(0,'RTU')

    baudrate_L = Label(frame2,text="傳輸速率",font="Keiu 16")
    baudrate_L.grid(row=2,column=0,pady=5)
    baudrate_E = Entry(frame2,width=8,font="Keiu 16")
    baudrate_E.grid(row=2,column=1,pady=5)
    baudrate_E.insert(0,'9600')

    bytesize_L = Label(frame2,text="字節大小",font="Keiu 16")
    bytesize_L.grid(row=3,column=0,pady=5)
    bytesize_E = Entry(frame2,width=8,font="Keiu 16")
    bytesize_E.grid(row=3,column=1,pady=5)
    bytesize_E.insert(0,'8')

    parity_L = Label(frame2,text="校驗",font="Keiu 16")
    parity_L.grid(row=4,column=0,pady=5)
    parity_E = Entry(frame2,width=8,font="Keiu 16")
    parity_E.grid(row=4,column=1,pady=5)
    parity_E.insert(0,'EVEN')

    stopbits_L = Label(frame2,text="停止位元",font="Keiu 16")
    stopbits_L.grid(row=5,column=0,pady=5)
    stopbits_E = Entry(frame2,width=8,font="Keiu 16")
    stopbits_E.grid(row=5,column=1,pady=5)
    stopbits_E.insert(0,'1')

    timeout_L = Label(frame2,text="超時",font="Keiu 16")
    timeout_L.grid(row=6,column=0,pady=5)
    timeout_E = Entry(frame2,width=8,font="Keiu 16")
    timeout_E.grid(row=6,column=1,pady=5)
    timeout_E.insert(0,'1')

    #-----豆溫------------------------------------------------------------------------
    bt_slaveaddress_L = Label(frame2,text="豆溫從動裝置", foreground='red',font="Keiu 16")
    bt_slaveaddress_L.grid(row=0,column=3,padx=10,pady=5)
    bt_slaveaddress_E = Entry(frame2,width=8,font="Keiu 16")
    bt_slaveaddress_E.grid(row=0,column=4,pady=5)
    bt_slaveaddress_E.insert(0,'1')

    bt_register_PV_L = Label(frame2,text="豆溫PV位址", foreground='red',font="Keiu 16")
    bt_register_PV_L.grid(row=1,column=3,padx=10,pady=5)
    bt_register_PV_E = Entry(frame2,width=8,font="Keiu 16")
    bt_register_PV_E.grid(row=1,column=4,pady=5)
    bt_register_PV_E.insert(0,'8192')#omron:H2000=8192    	Delta:18176

    bt_register_SV_L = Label(frame2,text="豆溫SV位址", foreground='red',font="Keiu 16")
    bt_register_SV_L.grid(row=2,column=3,padx=10,pady=5)
    bt_register_SV_E = Entry(frame2,width=8,font="Keiu 16")
    bt_register_SV_E.grid(row=2,column=4,pady=5)
    bt_register_SV_E.insert(0,'8451')#omron:H2003=8451       Delta:18177

    #-----環境溫------------------------------------------------------------------------
    et_slaveaddress_L = Label(frame2,text="環境溫從動裝置", foreground='blue',font="Keiu 16")
    et_slaveaddress_L.grid(row=4,column=3,padx=10,pady=5)
    et_slaveaddress_E = Entry(frame2,width=8,font="Keiu 16")
    et_slaveaddress_E.grid(row=4,column=4,pady=5)
    et_slaveaddress_E.insert(0,'2')

    et_register_PV_L = Label(frame2,text="環境溫PV位址", foreground='blue',font="Keiu 16")
    et_register_PV_L.grid(row=5,column=3,padx=10,pady=5)
    et_register_PV_E = Entry(frame2,width=8,font="Keiu 16")
    et_register_PV_E.grid(row=5,column=4,pady=5)
    et_register_PV_E.insert(0,'18176')#omron:H2000=8192    	Delta:18176

    et_register_SV_L = Label(frame2,text="環境溫SV位址", foreground='blue',font="Keiu 16")
    et_register_SV_L.grid(row=6,column=3,padx=10,pady=5)
    et_register_SV_E = Entry(frame2,width=8,font="Keiu 16")
    et_register_SV_E.grid(row=6,column=4,pady=5)
    et_register_SV_E.insert(0,'18177')#omron:H2003=8451       Delta:18177

    #-----入風溫------------------------------------------------------------------------
    entry_slaveaddress_L = Label(frame2,text="入風溫從動裝置", foreground='green',font="Keiu 16")
    entry_slaveaddress_L.grid(row=0,column=5,padx=10,pady=5)
    entry_slaveaddress_E = Entry(frame2,width=8,font="Keiu 16")
    entry_slaveaddress_E.grid(row=0,column=6,pady=5)
    entry_slaveaddress_E.insert(0,'0')

    entry_register_PV_L = Label(frame2,text="入風溫PV位址", foreground='green',font="Keiu 16")
    entry_register_PV_L.grid(row=1,column=5,padx=10,pady=5)
    entry_register_PV_E = Entry(frame2,width=8,font="Keiu 16")
    entry_register_PV_E.grid(row=1,column=6,pady=5)
    entry_register_PV_E.insert(0,'18176')#omron:H2000=8192    	Delta:18176

    entry_register_SV_L = Label(frame2,text="入風溫SV位址", foreground='green',font="Keiu 16")
    entry_register_SV_L.grid(row=2,column=5,padx=10,pady=5)
    entry_register_SV_E = Entry(frame2,width=8,font="Keiu 16")
    entry_register_SV_E.grid(row=2,column=6,pady=5)
    entry_register_SV_E.insert(0,'18177')#omron:H2003=8451       Delta:18177

    btn_1 = Button(frame2,text='確認設定', style='W1.TButton',command=lambda:notebook.select(0))#root_t.destroy
    btn_1.grid(row=10,column=1,padx=5,pady=5)
    
    btn_2 = Button(frame2,text="參數測試", style='W1.TButton',command=lambda:argument_setup(5))
    btn_2.grid(row=10,column=3,padx=5,pady=5)

    btn_3 = Button(frame2,text='重新選擇', style='W1.TButton',command=lambda:notebook.select(2))#root_t.destroy
    btn_3.grid(row=10,column=4,padx=5,pady=5)

    line = ttk.Separator(frame2, orient=VERTICAL, style="Line.TSeparator").grid(column=8, row=0, rowspan=14,padx=5,pady=5, sticky='ns')

    btn_4 = Button(frame2,text='選擇檔案', style='W1.TButton',command=lambda:load_file_data(3))
    btn_4.grid(row=0,column=10,padx=5,pady=5)

    select_file_E = Entry(frame2,width=20,font="Keiu 14")
    select_file_E.grid(row=0,column=12,padx=5,pady=5)

    #Logo插圖
    frame2_img_src = Image.open('roast_machine.jpg')
    frame2_img_src=frame2_img_src.resize((int(frame2_img_src.size[0]/2),int(frame2_img_src.size[1]/2)))
    frame2_img_tk = ImageTk.PhotoImage(frame2_img_src)
    frame2_img = Canvas(frame2,width=frame2_img_src.size[0],height=frame2_img_src.size[1])#width=img.size[0],height=img.size[1]
    frame2_img.create_image(0, 0, anchor=NW, image=frame2_img_tk)#tk_img'nw'
    frame2_img.grid(row=2,column=10,rowspan=10,columnspan=3,pady=5,padx=5)#,rowspan=15,columnspan=2

    #*************** 第三個視窗-系統選項 ***************
    frame3 = Frame(root_t, relief=RIDGE, borderwidth=2)
    frame3.pack(side=TOP, fill='both', ipadx="1c", ipady="1c", expand=1)

    #-------控制器通用參數--------------------------------------------------------------------------
    port_L = Label(frame3,text="通信端口",font="Keiu 16")
    port_L.grid(row=0,column=0,pady=5,padx=5)
    var_port = StringVar()    
    port_cb = Combobox(frame3,width=8,textvariable=var_port,font="Keiu 16")#
    port_cb["value"] = ('COM1','COM2','COM3','COM4','COM5','COM6','COM7','COM8','COM9')
    port_cb.current(4)
    port_cb.grid(row=0,column=1,pady=5,padx=5)

    mode_L = Label(frame3,text="通訊類型",font="Keiu 16")
    mode_L.grid(row=1,column=0,pady=5,padx=5)
    var_mode = StringVar()     
    mode_cb = Combobox(frame3,width=8,textvariable=var_mode,font="Keiu 16")#
    mode_cb["value"] = ('RTU','ASCII')
    mode_cb.current(0)
    mode_cb.grid(row=1,column=1,pady=5,padx=5)

    baudrate_L = Label(frame3,text="傳輸速率",font="Keiu 16")
    baudrate_L.grid(row=2,column=0,pady=5,padx=5)
    var_baudrate = StringVar()     
    baudrate_cb = Combobox(frame3,width=8,textvariable=var_baudrate,font="Keiu 16")#
    baudrate_cb["value"] = ('1200','2400','4800','9600','19200','38400','57600','115200')
    baudrate_cb.current(3)
    baudrate_cb.grid(row=2,column=1,pady=5,padx=5)

    bytesize_L = Label(frame3,text="字節大小",font="Keiu 16")
    bytesize_L.grid(row=3,column=0,pady=5,padx=5)
    var_bytesize = StringVar()     
    bytesize_cb = Combobox(frame3,width=8,textvariable=var_bytesize,font="Keiu 16")#
    bytesize_cb["value"] = ('7','8')
    bytesize_cb.current(1)
    bytesize_cb.grid(row=3,column=1,pady=5,padx=5)

    parity_L = Label(frame3,text="校驗",font="Keiu 16")
    parity_L.grid(row=4,column=0,pady=5,padx=5)
    var_parity = StringVar()     
    parity_cb = Combobox(frame3,width=8,textvariable=var_parity,font="Keiu 16")#
    parity_cb["value"] = ('NONE','EVEN','ODD','MARK','SPACE')
    parity_cb.current(1)
    parity_cb.grid(row=4,column=1,pady=5,padx=5)

    stopbits_L = Label(frame3,text="停止位元",font="Keiu 16")
    stopbits_L.grid(row=5,column=0,pady=5,padx=5)
    var_stopbits = StringVar()     
    stopbits_cb = Combobox(frame3,width=8,textvariable=var_stopbits,font="Keiu 16")#
    stopbits_cb["value"] = ('1','2')
    stopbits_cb.current(0)
    stopbits_cb.grid(row=5,column=1,pady=5,padx=5)

    timeout_L = Label(frame3,text="超時",font="Keiu 16")
    timeout_L.grid(row=6,column=0,pady=5,padx=5)
    var_timeout = StringVar()     
    timeout_cb = Combobox(frame3,width=8,textvariable=var_timeout,font="Keiu 16")#
    timeout_cb["value"] = ('0.2','0.5','1','2')
    timeout_cb.current(2)
    timeout_cb.grid(row=6,column=1,pady=5,padx=5)
    #-------豆溫--------------------------------------------------------------------------
    bt_slaveaddress_L = Label(frame3, foreground='red',text="豆溫從動裝置",font="Keiu 16")
    bt_slaveaddress_L.grid(row=0,column=2,pady=5,padx=5)
    bt_var_slaveaddress = StringVar()     
    bt_slaveaddress_cb = Combobox(frame3,width=8,textvariable=bt_var_slaveaddress,font="Keiu 16")#
    bt_slaveaddress_cb["value"] = ('0','1','2','3','4','5','6','7','8')
    bt_slaveaddress_cb.current(1)
    bt_slaveaddress_cb.grid(row=0,column=3,pady=5,padx=5)

    bt_register_PV_L = Label(frame3, foreground='red',text="豆溫PV位址",font="Keiu 16")
    bt_register_PV_L.grid(row=1,column=2,pady=5,padx=5)
    bt_var_register_PV = StringVar()     
    bt_register_PV_cb = Combobox(frame3,width=8,textvariable=bt_var_register_PV,font="Keiu 16")#
    bt_register_PV_cb["value"] = ('8192','18176')#Omron , Delta
    bt_register_PV_cb.current(0)
    bt_register_PV_cb.grid(row=1,column=3,pady=5,padx=5)

    bt_register_SV_L = Label(frame3, foreground='red',text="豆溫SV位址",font="Keiu 16")
    bt_register_SV_L.grid(row=2,column=2,pady=5,padx=5)
    bt_var_register_SV = StringVar()     
    bt_register_SV_cb = Combobox(frame3,width=8,textvariable=bt_var_register_SV,font="Keiu 16")#
    bt_register_SV_cb["value"] = ('8451','18177')#Omron , Delta
    bt_register_SV_cb.current(0)
    bt_register_SV_cb.grid(row=2,column=3,pady=5,padx=5)

    #-------環境溫--------------------------------------------------------------------------
    et_slaveaddress_L = Label(frame3, foreground='blue',text="環境溫從動裝置",font="Keiu 16")
    et_slaveaddress_L.grid(row=4,column=2,pady=5,padx=5)
    et_var_slaveaddress = StringVar()     
    et_slaveaddress_cb = Combobox(frame3,width=8,textvariable=et_var_slaveaddress,font="Keiu 16")#
    et_slaveaddress_cb["value"] = ('0','1','2','3','4','5','6','7','8')
    et_slaveaddress_cb.current(2)
    et_slaveaddress_cb.grid(row=4,column=3,pady=5,padx=5)

    et_register_PV_L = Label(frame3, foreground='blue',text="環境溫PV位址",font="Keiu 16")
    et_register_PV_L.grid(row=5,column=2,pady=5,padx=5)
    et_var_register_PV = StringVar()     
    et_register_PV_cb = Combobox(frame3,width=8,textvariable=et_var_register_PV,font="Keiu 16")#
    et_register_PV_cb["value"] = ('8192','18176')#Omron , Delta
    et_register_PV_cb.current(1)
    et_register_PV_cb.grid(row=5,column=3,pady=5,padx=5)

    et_register_SV_L = Label(frame3, foreground='blue',text="環境溫SV位址",font="Keiu 16")
    et_register_SV_L.grid(row=6,column=2,pady=5,padx=5)
    et_var_register_SV = StringVar()     
    et_register_SV_cb = Combobox(frame3,width=8,textvariable=et_var_register_SV,font="Keiu 16")#
    et_register_SV_cb["value"] = ('8451','18177')#Omron , Delta
    et_register_SV_cb.current(1)
    et_register_SV_cb.grid(row=6,column=3,pady=5,padx=5)
    #-------入風溫--------------------------------------------------------------------------
    entry_slaveaddress_L = Label(frame3, foreground='green',text="入風溫從動裝置",font="Keiu 16")
    entry_slaveaddress_L.grid(row=0,column=4,pady=5,padx=5)
    entry_var_slaveaddress = StringVar()     
    entry_slaveaddress_cb = Combobox(frame3,width=8,textvariable = entry_var_slaveaddress,font="Keiu 16")#
    entry_slaveaddress_cb["value"] = ('0','1','2','3','4','5','6','7','8')
    entry_slaveaddress_cb.current(0)
    entry_slaveaddress_cb.grid(row=0,column=5,pady=5,padx=5)

    entry_register_PV_L = Label(frame3, foreground='green',text="入風溫PV位址",font="Keiu 16")
    entry_register_PV_L.grid(row=1,column=4,pady=5,padx=5)
    entry_var_register_PV = StringVar()     
    entry_register_PV_cb = Combobox(frame3,width=8,textvariable=entry_var_register_PV,font="Keiu 16")#
    entry_register_PV_cb["value"] = ('8192','18176')#Omron , Delta
    entry_register_PV_cb.current(0)
    entry_register_PV_cb.grid(row=1,column=5,pady=5,padx=5)

    entry_register_SV_L = Label(frame3, foreground='green',text="入風溫SV位址",font="Keiu 16")
    entry_register_SV_L.grid(row=2,column=4,pady=5,padx=5)
    entry_var_register_SV = StringVar()     
    entry_register_SV_cb = Combobox(frame3,width=8,textvariable=entry_var_register_SV,font="Keiu 16")#
    entry_register_SV_cb["value"] = ('8451','18177')#Omron , Delta
    entry_register_SV_cb.current(0)
    entry_register_SV_cb.grid(row=2,column=5,pady=5,padx=5)

    Button(frame3,text="連線參數", style='W1.TButton',command=lambda:argument_setup(0)).grid(row=8,column=0,columnspan=5,padx=5,pady=5)
    #Button(frame3,text="參數測試", style='W1.TButton',command=lambda:argument_setup(5)).grid(row=8,column=3,padx=5,pady=5)

    #分隔線
    line_3_1 = ttk.Separator(frame3, orient=VERTICAL, style="Line.TSeparator").grid(row=0, rowspan=14, column=6, sticky='ns')

    #*****----- 烘焙階段 溫度起迄設定-----*****

    rostep_L = Label(frame3,width=10,text="烘焙階段", foreground='red',font="Keiu 16")#烘焙階段
    rostep_L.grid(row=0,column=7,columnspan=3,padx=5,pady=5)
    rostep_start_L = Label(frame3,width=5,text="開始", foreground='blue',font="Keiu 16")#開始
    rostep_start_L.grid(row=1,column=8,padx=5,pady=5)
    rostep_end_L = Label(frame3,width=5,text="結束", foreground='blue',font="Keiu 16")#結束
    rostep_end_L.grid(row=1,column=9,padx=5,pady=5)

    rostep_dry_L = Label(frame3,width=10,text="脫水階段", foreground='green',font="Keiu 16")#脫水階段
    rostep_dry_L.grid(row=2,column=7,padx=5,pady=5)
    rostep_dry_start_E = Entry(frame3,width=5,font="Keiu 16")
    rostep_dry_start_E.grid(row=2,column=8,padx=5,pady=5)
    rostep_dry_start_E.insert(0,'110')#內定脫水起始溫度 
    rostep_dry_end_E = Entry(frame3,width=5,font="Keiu 16")
    rostep_dry_end_E.grid(row=2,column=9,padx=5,pady=5)
    rostep_dry_end_E.insert(0,'150')#內定脫水結束溫度 
       
    rostep_maillard_L = Label(frame3,width=10,text="梅納階段", foreground='#ff8000',font="Keiu 16")#梅納階段
    rostep_maillard_L.grid(row=3,column=7,padx=5,pady=5)
    rostep_maillard_start_E = Entry(frame3,width=5,font="Keiu 16")
    rostep_maillard_start_E.grid(row=3,column=8,padx=5,pady=5)
    rostep_maillard_start_E.insert(0,'150')#內定梅納階段起始溫度 
    rostep_maillard_end_E = Entry(frame3,width=5,font="Keiu 16")
    rostep_maillard_end_E.grid(row=3,column=9,padx=5,pady=5)
    rostep_maillard_end_E.insert(0,'190')#內定梅納階段結束溫度 

    rostep_development_L = Label(frame3,width=10,text="完成階段", foreground='brown',font="Keiu 16")#完成階段
    rostep_development_L.grid(row=4,column=7,padx=5,pady=5)
    rostep_development_start_E = Entry(frame3,width=5,font="Keiu 16")
    rostep_development_start_E.grid(row=4,column=8,padx=5,pady=5)
    rostep_development_start_E.insert(0,'190')#內定完成階段起始溫度 
    rostep_development_end_E = Entry(frame3,width=5,font="Keiu 16")
    rostep_development_end_E.grid(row=4,column=9,padx=5,pady=5)
    rostep_development_end_E.insert(0,'230')#內定完成階段結束溫度 
    Button(frame3,text="重設階段",style='W.TButton', command=lambda:step_change()).grid(row=8,column=7,columnspan=3,padx=5,pady=5)

    #*****----- 烘焙階段 End -----*****

    #分隔線
    line_3_3 = ttk.Separator(frame3, orient=VERTICAL, style="Line.TSeparator").grid(row=0, rowspan=14,column=14, sticky='ns')

    #橫分隔線
    line_3_5 = ttk.Separator(frame3, orient=HORIZONTAL, style="Line.TSeparator").grid(row=18,column=0, columnspan=20, sticky='ew')#

    Label(frame3,text="儲存烘豆參數",font="Keiu 16").grid(row=20,column=0,pady=5,padx=5)
    Label(frame3,text="檔案名稱",foreground='blue',font="Keiu 14").grid(row=20,column=1,pady=5,padx=5)
    equipment_name = Entry(frame3,font="Keiu 16")
    equipment_name.grid(row=20,column=2,columnspan=1,pady=5,padx=5)
    equipment_name.insert(0,'arg_test')

    Button(frame3,text="參數存檔", style='W1.TButton',command=lambda:argument_setup(1)).grid(row=20,column=3,padx=5,pady=5)
    #橫分隔線
    line_3_6 = ttk.Separator(frame3, orient=HORIZONTAL, style="Line.TSeparator").grid(row=22,column=0, columnspan=20, sticky='ew')#

    frame3_modu() #顯示不同溫控器的暫存器位址
    #分隔線
    line_3_3 = ttk.Separator(frame3, orient=VERTICAL, style="Line.TSeparator").grid(row=22, rowspan=14,column=4, sticky='ns')

    #*****----- 全息烘焙 T0、T1、T2 設定溫度 -----*****
    all_rost_L = Label(frame3,width=10,text="全息烘焙", foreground='green',font="Keiu 16")#全息烘焙
    all_rost_L.grid(row=23,column=5,columnspan=6,padx=5,pady=5)

    all_rost_temps_L = Label(frame3,width=5,text="min", foreground='blue',font="Keiu 16")#開始
    all_rost_temps_L.grid(row=25,column=6,padx=5,pady=5)
    all_rost_tempe_L = Label(frame3,width=5,text="max", foreground='blue',font="Keiu 16")#結束
    all_rost_tempe_L.grid(row=25,column=7,padx=5,pady=5)

    all_rost_times_L = Label(frame3,width=6,text="start_T", foreground='blue',font="Keiu 16")#開始
    all_rost_times_L.grid(row=25,column=8,padx=5,pady=5)
    all_rost_timee_L = Label(frame3,width=6,text="end_T", foreground='blue',font="Keiu 16")#結束
    all_rost_timee_L.grid(row=25,column=9,padx=5,pady=5)

    tp_temp_L = Label(frame3,text="回溫點溫度°C",font="Keiu 16")#T0 點溫度
    tp_temp_L.grid(row=27,column=5,padx=5,pady=5)
    tp_temp_E_min = Entry(frame3,width=5,font="Keiu 16")
    tp_temp_E_min.grid(row=27,column=6,padx=5,pady=5)
    tp_temp_E_min.insert(0,'70')
    tp_temp_E_max = Entry(frame3,width=5,font="Keiu 16")
    tp_temp_E_max.grid(row=27,column=7,padx=5,pady=5)
    tp_temp_E_max.insert(0,'90')

    tp_times_E = Entry(frame3,width=5,font="Keiu 16")
    tp_times_E.grid(row=27,column=8,padx=5,pady=5)
    tp_times_E.insert(0,'0:30')
    tp_timee_E = Entry(frame3,width=5,font="Keiu 16")
    tp_timee_E.grid(row=27,column=9,padx=5,pady=5)
    tp_timee_E.insert(0,'1:30')

    t0_temp_L = Label(frame3,text="T0 點溫度°C",font="Keiu 16")#T0 點溫度
    t0_temp_L.grid(row=29,column=5,padx=5,pady=5)
    t0_temp_E_min = Entry(frame3,width=5,font="Keiu 16")
    t0_temp_E_min.grid(row=29,column=6,padx=5,pady=5)
    t0_temp_E_min.insert(0,'90')
    t0_temp_E_max = Entry(frame3,width=5,font="Keiu 16")
    t0_temp_E_max.grid(row=29,column=7,padx=5,pady=5)
    t0_temp_E_max.insert(0,'100')

    t0_times_E = Entry(frame3,width=5,font="Keiu 16")
    t0_times_E.grid(row=29,column=8,padx=5,pady=5)
    t0_times_E.insert(0,'1:30')
    t0_timee_E = Entry(frame3,width=5,font="Keiu 16")
    t0_timee_E.grid(row=29,column=9,padx=5,pady=5)
    t0_timee_E.insert(0,'2:00')    

    t1_temp_L = Label(frame3,text="T1 點溫度°C",font="Keiu 16")#T1 點溫度
    t1_temp_L.grid(row=31,column=5,padx=5,pady=5)
    t1_temp_E_min = Entry(frame3,width=5,font="Keiu 16")
    t1_temp_E_min.grid(row=31,column=6,padx=5,pady=5)
    t1_temp_E_min.insert(0,'110')
    t1_temp_E_max = Entry(frame3,width=5,font="Keiu 16")
    t1_temp_E_max.grid(row=31,column=7,padx=5,pady=5)
    t1_temp_E_max.insert(0,'120')

    t1_times_E = Entry(frame3,width=5,font="Keiu 16")
    t1_times_E.grid(row=31,column=8,padx=5,pady=5)
    t1_times_E.insert(0,'2:30')
    t1_timee_E = Entry(frame3,width=5,font="Keiu 16")
    t1_timee_E.grid(row=31,column=9,padx=5,pady=5)
    t1_timee_E.insert(0,'3:30')    

    t2_temp_L = Label(frame3,text="T2 點溫度°C",font="Keiu 16")#T2 點溫度
    t2_temp_L.grid(row=33,column=5,padx=5,pady=5)
    t2_temp_E_min = Entry(frame3,width=5,font="Keiu 16")
    t2_temp_E_min.grid(row=33,column=6,padx=5,pady=5)
    t2_temp_E_min.insert(0,'170')
    t2_temp_E_max = Entry(frame3,width=5,font="Keiu 16")
    t2_temp_E_max.grid(row=33,column=7,padx=5,pady=5)
    t2_temp_E_max.insert(0,'195')

    t2_times_E = Entry(frame3,width=5,font="Keiu 16")
    t2_times_E.grid(row=33,column=8,padx=5,pady=5)
    t2_times_E.insert(0,'1:00')
    t2_timee_E = Entry(frame3,width=5,font="Keiu 16")
    t2_timee_E.grid(row=33,column=9,padx=5,pady=5)
    t2_timee_E.insert(0,'2:30')    

    #*****----- 全息烘焙 T0、T1、T2 設定溫度 End -----*****
    Button(frame3,text="回BT & RoR", style='W.TButton',command=lambda:notebook.select(0)).grid(row=35,column=1,padx=5,pady=5)    
    Button(frame3,text="結束程式", style='W.TButton',command=root_t.destroy).grid(row=35,column=3,padx=5,pady=5)

    #*************** 第四個視窗-烘豆手法 ***************
    frame4 = Frame (root_t, relief=GROOVE, borderwidth=2)
    frame4.pack(side=BOTTOM, fill='both', ipadx="1c", ipady="1c", expand=1)
    frame4_modu()
    #*************** 第四個視窗-烘豆手法 End ***************
    
    #*************** 第五個視窗-烘豆紀錄表 ***************
    frame5 = Frame (root_t, relief=GROOVE, borderwidth=2)
    frame5.pack(side=TOP, fill='both', ipadx="1c", ipady="1c", expand=1)
    #********************* 資料記錄表格 ******************
    style_value = ttk.Style()
    style_value.configure("Treeview",background='lightgreen',foreground='blue',rowheight=30,font=("keui",14))#,foreground='blue'
    tree = Treeview(frame5,column=("item","BT (°C)","BT RoR","ET (°C)","ET RoR","時間(分)","事件"), show = 'headings', height=30, selectmode='browse')#

    yscrollbar =Scrollbar(tree)
    yscrollbar.config( orient="vertical",command=tree.yview)    
    yscrollbar.pack(side='right',fill='y')

    tree.configure(yscrollcommand=yscrollbar.set)

    tree.heading("#1",text='item')
    tree.heading("#2",text='BT (°C)')
    tree.heading("#3",text='BT RoR')
    tree.heading("#4",text='ET (°C)')
    tree.heading("#5",text='ET RoR')
    tree.heading("#6",text='時間(分)')
    tree.heading("#7",text='事件')

    tree.column("#1",anchor=CENTER,width=30)
    tree.column("#2",anchor=CENTER,width=80)
    tree.column("#3",anchor=CENTER,width=80)
    tree.column("#4",anchor=CENTER,width=80)
    tree.column("#5",anchor=CENTER,width=80)
    tree.column("#6",anchor=CENTER,width=80)
    tree.column("#7",anchor=CENTER,width=80)

    tree.grid(row=0, rowspan=10, column=10,columnspan=8, padx=5, pady=5, sticky='nsew')
    #******************** 資料記錄表格 END *******************
    filename_L = Label(frame5,text="檔案名稱",font="Keiu 16")#記錄檔名
    filename_L.grid(row=0,column=1,padx=5,pady=5)
    filename_E = Entry(frame5,width=20,font="Keiu 16")
    filename_E.grid(row=0,column=2,padx=5,pady=5)
    default_name = 'Akro-'+' '+'-'+str(datetime.date.today())+'-Test'
    filename_E.insert(0,default_name)#內定檔名

    roastdate_L = Label(frame5,text="烘焙日期",font="Keiu 16")#烘焙日期
    roastdate_L.grid(row=1,column=1,padx=5,pady=5)
    roastdate_E = Entry(frame5,width=20,font="Keiu 16")
    roastdate_E.grid(row=1,column=2,padx=5,pady=5)
    roastdate_E.insert(0,datetime.date.today())#內定日期

    prodname_L = Label(frame5,text="產品名稱",font="Keiu 16")#記錄產品名稱
    prodname_L.grid(row=2,column=1,padx=5,pady=5)
    prodname_E = Entry(frame5,width=20,font="Keiu 16")
    prodname_E.grid(row=2,column=2,padx=5,pady=5)
    prodname_E.insert(0,'SJ Coffee')#內定產品名稱

    #********** 跳脫繪圖狀態的隱藏開關 **********
    greenbean_inf_L = Label(frame5,text="生豆資訊", foreground='green',font="Keiu 16")#生豆資訊
    greenbean_inf_L.grid(row=3,column=0,columnspan=3,padx=5,pady=5)
    #********** 跳脫繪圖狀態的隱藏開關 End **********

    gb_method_L = Label(frame5,text="生豆處理法", foreground='green',font="Keiu 16")#生豆處理法
    gb_method_L.grid(row=4,column=1,padx=5,pady=5)
    gb_method_E = Entry(frame5,width=20,font="Keiu 16")
    gb_method_E.grid(row=4,column=2,padx=5,pady=5)
    #gb_method_E.insert(0,'水洗')#內定生豆處理法 
       
    gb_name_L = Label(frame5,text="生豆名稱", foreground='green',font="Keiu 16")#生豆名稱
    gb_name_L.grid(row=5,column=1,padx=5,pady=5)
    gb_name_E = Entry(frame5,width=20,font="Keiu 16")
    gb_name_E.grid(row=5,column=2,padx=5,pady=5)
    #gb_name_E.insert(0,'耶加雪菲')#內定生豆名稱

    gb_moisture_L = Label(frame5,text="水分含量", foreground='green',font="Keiu 16")#水分含量
    gb_moisture_L.grid(row=6,column=1,padx=5,pady=5)
    gb_moisture_E = Entry(frame5,width=20,font="Keiu 16")
    gb_moisture_E.grid(row=6,column=2,padx=5,pady=5)
    #gb_moisture_E.insert(0,'11.2%')#內定水分含量

    gb_density_L = Label(frame5,text="生豆密度", foreground='green',font="Keiu 16")#生豆密度
    gb_density_L.grid(row=7,column=1,padx=5,pady=5)
    gb_density_E = Entry(frame5,width=20,font="Keiu 16")
    gb_density_E.grid(row=7,column=2,padx=5,pady=5)
    #gb_density_E.insert(0,'834 g/L')#內定生豆密度

    gb_weight_L = Label(frame5,text="批次重量", foreground='green',font="Keiu 16")#批次重量
    gb_weight_L.grid(row=8,column=1,padx=5,pady=5)
    gb_weight_E = Entry(frame5,width=20,font="Keiu 16")
    gb_weight_E.grid(row=8,column=2,padx=5,pady=5)
    #gb_weight_E.insert(0,'1200 G')#內定批次重量

    equipment_L = Label(frame5,text="烘豆機資訊", foreground='blue',font="Keiu 16")#烘豆機資訊
    equipment_L.grid(row=10,column=0,columnspan=3,padx=5,pady=5)

    machine_name_L = Label(frame5,text="設備名稱", foreground='blue',font="Keiu 16")#設備名稱
    machine_name_L.grid(row=11,column=1,padx=5,pady=5)
    machine_name_E = Entry(frame5,width=20,font="Keiu 16")
    machine_name_E.grid(row=11,column=2,padx=5,pady=5)
    machine_name_E.insert(0,'亦安機械 AKRO')#內定設備名稱 
       
    content_L = Label(frame5,text="設備容量", foreground='blue',font="Keiu 16")#設備容量
    content_L.grid(row=12,column=1,padx=5,pady=5)
    content_E = Entry(frame5,width=20,font="Keiu 16")
    content_E.grid(row=12,column=2,padx=5,pady=5)
    content_E.insert(0,'1.2Kg')#內定設備容量

    energy_L = Label(frame5,text="使用能源", foreground='blue',font="Keiu 16")#使用能源
    energy_L.grid(row=13,column=1,padx=5,pady=5)
    energy_E = Entry(frame5,width=20,font="Keiu 16")
    energy_E.grid(row=13,column=2,padx=5,pady=5)
    energy_E.insert(0,'LPG')#內定使用能源

    operator_L = Label(frame5,text="操作人員", foreground='blue',font="Keiu 16")#操作人員
    operator_L.grid(row=14,column=1,padx=5,pady=5)
    operator_E = Entry(frame5,width=20,font="Keiu 16")
    operator_E.grid(row=14,column=2,padx=5,pady=5)
    operator_E.insert(0,'W J Lin')#內定操作人員

    weather_L = Label(frame5,text="天氣溫度", foreground='blue',font="Keiu 16")#天氣溫度
    weather_L.grid(row=15,column=1,padx=5,pady=5)
    weather_E = Entry(frame5,width=20,font="Keiu 16")
    weather_E.grid(row=15,column=2,padx=5,pady=5)
    weather_E.insert(0,'晴天 32.5 °C')#內定天氣溫度

    #********** 熟豆資訊 **********
    coffeebean_inf_L = Label(frame5,text="熟豆資訊", foreground='brown',font="Keiu 16")#熟豆資訊
    coffeebean_inf_L.grid(row=3,column=20,columnspan=3,padx=5,pady=5)

    cb_weight_L = Label(frame5,text="熟豆重量", foreground='brown',font="Keiu 16")#熟豆重量
    cb_weight_L.grid(row=4,column=20,padx=5,pady=5)
    cb_weight_E = Entry(frame5,width=20,font="Keiu 16")
    cb_weight_E.grid(row=4,column=21,padx=5,pady=5)
    #cb_area_E.insert(0,'485')#內定熟豆重量

    loss_weight_L = Label(frame5,text="失重率", foreground='brown',font="Keiu 16")#失重率
    loss_weight_L.grid(row=5,column=20,padx=5,pady=5)
    loss_weight_E = Entry(frame5,width=20,font="Keiu 16")
    loss_weight_E.grid(row=5,column=21,padx=5,pady=5)
    #loss_weight_E.insert(0,'13.75%')#內定失重率

    cb_barg_L = Label(frame5,text="豆色焙度", foreground='brown',font="Keiu 16")#豆色焙度
    cb_barg_L.grid(row=6,column=20,padx=5,pady=5)
    cb_barg_E = Entry(frame5,width=20,font="Keiu 16")
    cb_barg_E.grid(row=6,column=21,padx=5,pady=5)
    #cb_barg_E.insert(0,'75')#內定豆色焙度 

    cb_parg_L = Label(frame5,text="粉色焙度", foreground='brown',font="Keiu 16")#粉色焙度
    cb_parg_L.grid(row=7,column=20,padx=5,pady=5)
    cb_parg_E = Entry(frame5,width=20,font="Keiu 16")
    cb_parg_E.grid(row=7,column=21,padx=5,pady=5)
    #cb_parg_E.insert(0,'88')#內定粉色焙度 

    cb_cupping_L = Label(frame5,text="杯測風味", foreground='brown',font="Keiu 16")#杯測風味
    cb_cupping_L.grid(row=8,column=20,padx=5,pady=5)
    cb_cupping_E = Entry(frame5,width=20,font="Keiu 16")
    cb_cupping_E.grid(row=8,column=21,padx=5,pady=5)
    #gb_cupping_E.insert(0,'柑橘調性')#內定杯測風味 

    Button(frame5,text="選擇已存資料",style='W3.TButton', command=lambda:load_file_data(1)).grid(row=11,column=10,padx=10,pady=10)#,columnspan=17
    Button(frame5,text="讀Artisan檔案",style='W3.TButton', command=lambda:load_file_data(6)).grid(row=11,column=12,padx=10,pady=10)
    Button(frame5,text="產生紀錄表",state='',style='W3.TButton', command=lambda:load_file_data(2)).grid(row=11,column=14,padx=10,pady=10)
    Button(frame5,text="設備參數位址",state=DISABLED,style='W3.TButton', command=lambda:notebook.select(3)).grid(row=11,column=16,padx=10,pady=10)
    #*************** 第五個視窗-烘豆紀錄表 End ***************

    #*************** 第六個視窗-全息烘焙 ***************
    frame6 = Frame (root_t, relief=GROOVE, borderwidth=2)
    frame6.pack(side=TOP, fill='both', ipadx="1c", ipady="1c", expand=1)

    canvas_ss =Canvas(frame6, width=940, height=460, bg='white')
    canvas_ss.grid(row=2, rowspan=13 ,column=2, columnspan=17, padx=5, pady=5)

    fc_agtron_L = Label(frame6,text="FC 豆表 Agtron",font="Keiu 16")#一爆時的艾格狀數
    fc_agtron_L.grid(row=15,column=1,padx=5,pady=5)
    fc_agtron_E = Entry(frame6)#,style='W.TButton'
    fc_agtron_E.grid(row=15,column=2)
    fc_agtron_E.insert(0,'100')

    agtron_L = Label(frame6,text="豆表 Agtron 預估值",font="Keiu 16")#當下艾格狀數預估
    agtron_L.grid(row=15,column=4,padx=5,pady=5)
    agtron_L_S = Label(frame6,text="",font="Keiu 16")#FC 豆表 Agtron
    agtron_L_S.grid(row=15,column=5,padx=5,pady=5)

    mark_bt_ss_l = Label(frame6,text='BT', width = 6, background="white",foreground="red",font="Helvetica 20 bold").grid(row=16,column=1,pady=5,padx=5)
    bt_temp_ss_l = Label(frame6, width = 8, background="yellow",foreground="red",font="Helvetica 20 bold")
    bt_temp_ss_l.grid(row=16,column=2,pady=5,padx=5)

    mark_bt_ror_ss_l = Label(frame6,text='△BT', width = 6, background="white",foreground="green",font="Helvetica 20 bold").grid(row=17,column=1,pady=5,padx=5)
    bt_temp_ror_ss_l = Label(frame6, width = 8, background="yellow",foreground="green",font="Helvetica 20 bold")
    bt_temp_ror_ss_l.grid(row=17,column=2,pady=5,padx=5)

    #*************** 第六個視窗-全息烘焙 End ***************
    draw_panal()#產生座標軸
    draw_panal_ss(0)#產生座標軸

    notebook.add(frame1,text='繪圖 BT & ROR')
    notebook.add(frame2,text='通訊參數')
    notebook.add(frame3,text='參數選項')
    notebook.add(frame4,text='烘豆手法')
    notebook.add(frame5,text='烘豆紀錄表')
    notebook.add(frame6,text='全息烘焙')
    notebook.pack(padx=10,pady=10,fill=BOTH,expand=TRUE)
 
    root_t.mainloop()

####----------    測試用程式碼    ----------####
    '''
    #將TreeView中的資料寫入檔案中
    tree_item ='I001'
    with open( filename +'.rxt','w') as f:
        f.write(roastdate_E.get())  #烘焙日期
        f.write('\n')
        f.write(prodname_E.get())   #產品名稱
        f.write('\n')
        f.write(gb_area_E.get())   #生豆產地
        f.write('\n')
        f.write(gb_name_E.get())   #生豆名稱
        f.write('\n')
        f.write(gb_moisture_E.get())   #水分含量
        f.write('\n')
        f.write(gb_density_E.get())   #生豆密度
        f.write('\n')
        f.write(gb_weight_E.get())   #批次重量
        f.write('\n')
        f.write(machine_name_E.get())   #設備名稱
        f.write('\n')
        f.write(content_E.get())   #設備容量
        f.write('\n')
        f.write(energy_E.get())   #使用能源
        f.write('\n')
        f.write(operator_E.get())   #操作人員
        f.write('\n')
        f.write(weather_E.get())   #天氣溫度
        f.write('\n')

        while tree.exists(tree_item):#:next(tree.item)!= '',,tree.exists(tree_item)
            i += 1
            try:
                td_1,td_2,td_3,td_4,td_5,td_6 = tree.item(tree_item,option='values')
                #td_0:項次  td_1:豆溫  td_2:豆溫ROR  td_3:環境溫  td_4:環境溫ROR  td_5:烘豆時間  td_6:事件
                td_0 = i
                f.write(td_0+':'+td_1+':'+td_2+':'+td_3+':'+td_4+':'+td_5+':'+td_6+'\n')
                #print(tree.item(tree_item,option='values')[0],tree.item(tree_item,option='values')[1])    
                tree_item = tree.next(tree_item)
            except:
                print("存檔完成")
                break
    '''
    '''
    root_record = Tk()
    screenwidth = root_record.winfo_screenwidth()
    screenheight = root_record.winfo_screenheight()
    w_win = 1480
    h_win = 760
    x_offset = (screenwidth - w_win) / 2
    y_offset = ((screenheight - h_win) / 2)# - 30
    root_record.title("轉成Excel烘焙紀錄表")
    root_record.geometry("%dx%d+%d-%d" %(w_win,h_win,x_offset,y_offset))#1520
    root_record.configure(bg='lightblue')
    root_record.iconbitmap("coffee_logo.ico")

    root_record.mainloop()
    '''
####----------    測試用程式碼End    ----------####
